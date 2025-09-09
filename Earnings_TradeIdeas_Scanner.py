import FreeSimpleGUI as sg
import yfinance as yf
from datetime import datetime, timedelta
from scipy.interpolate import interp1d
import numpy as np
import pandas as pd
import threading
import time
import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import json

# =============================================================================
# CONFIGURATION - REALISTIC THRESHOLDS
# =============================================================================
VOLUME_THRESHOLD = 500000        # Reduced from 1,500,000 (more inclusive)
IV_RV_THRESHOLD = 0.9           # Reduced from 1.25 (more realistic) 
SLOPE_THRESHOLD = -0.001        # Relaxed from -0.00406 (less strict)

# API Keys and endpoints - PRIORITY ORDER
FINNHUB_API_KEY = "d2hfijpr01qon4ec0eu0d2hfijpr01qon4ec0eug"

# Base URLs
FINNHUB_BASE_URL = "https://finnhub.io/api/v1"

# =============================================================================
# DATA SOURCE CLASSES
# =============================================================================

class DataSource:
    """Base class for data sources"""
    
    def get_current_price(self, symbol):
        raise NotImplementedError
    
    def get_options_expirations(self, symbol):
        raise NotImplementedError
        
    def get_options_chain(self, symbol, expiration):
        raise NotImplementedError
        
    def get_price_history(self, symbol, period_days=90):
        raise NotImplementedError

class YahooFinanceSource(DataSource):
    """Yahoo Finance data source using yfinance - PRIMARY"""
    
    def __init__(self):
        self.name = "Yahoo Finance"
    
    def safe_call(self, func, max_retries=2, delay=2):
        """Wrapper for yfinance calls with retry logic"""
        for attempt in range(max_retries):
            try:
                result = func()
                return result
            except Exception as e:
                if attempt < max_retries - 1:
                    time.sleep(delay)
                else:
                    raise e
    
    def get_current_price(self, symbol):
        try:
            def get_price():
                ticker = yf.Ticker(symbol)
                hist = ticker.history(period='2d')
                if hist.empty:
                    raise ValueError("No price data")
                return float(hist['Close'].iloc[-1])
            
            return self.safe_call(get_price)
        except Exception as e:
            print(f"Yahoo Finance price error for {symbol}: {e}")
            return None
    
    def get_options_expirations(self, symbol):
        try:
            def get_expirations():
                ticker = yf.Ticker(symbol)
                options_dates = list(ticker.options)
                if not options_dates:
                    raise ValueError("No options data")
                return options_dates
            
            return self.safe_call(get_expirations)
        except Exception as e:
            print(f"Yahoo Finance options expiration error for {symbol}: {e}")
            return None
    
    def get_options_chain(self, symbol, expiration):
        try:
            def get_chain():
                ticker = yf.Ticker(symbol)
                chain = ticker.option_chain(expiration)
                if chain.calls.empty and chain.puts.empty:
                    raise ValueError("Empty options chain")
                return {
                    'calls': chain.calls,
                    'puts': chain.puts
                }
            
            return self.safe_call(get_chain)
        except Exception as e:
            print(f"Yahoo Finance options chain error for {symbol} {expiration}: {e}")
            return None
    
    def get_price_history(self, symbol, period_days=90):
        try:
            def get_history():
                ticker = yf.Ticker(symbol)
                hist = ticker.history(period=f"{period_days}d")
                if hist.empty:
                    raise ValueError("No price history")
                return hist
            
            return self.safe_call(get_history)
        except Exception as e:
            print(f"Yahoo Finance history error for {symbol}: {e}")
            return None

class FinnhubSource(DataSource):
    """Finnhub API data source - BACKUP"""
    
    def __init__(self, api_key):
        self.api_key = api_key
        self.name = "Finnhub"
        self.base_url = FINNHUB_BASE_URL
    
    def make_request(self, endpoint, params=None):
        """Make API request to Finnhub"""
        if params is None:
            params = {}
        params['token'] = self.api_key
        
        try:
            response = requests.get(f"{self.base_url}/{endpoint}", params=params, timeout=10)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            print(f"Finnhub API error: {e}")
            return None
    
    def get_current_price(self, symbol):
        try:
            data = self.make_request("quote", {"symbol": symbol})
            if data and 'c' in data:
                return float(data['c'])  # Current price
            return None
        except Exception as e:
            print(f"Finnhub price error for {symbol}: {e}")
            return None
    
    def get_options_expirations(self, symbol):
        """Generate standard monthly expirations"""
        try:
            expirations = []
            today = datetime.now().date()
            
            for month_offset in range(12):
                year = today.year
                month = today.month + month_offset
                
                if month > 12:
                    year += (month - 1) // 12
                    month = ((month - 1) % 12) + 1
                
                # Find 3rd Friday of the month
                first_day = datetime(year, month, 1).date()
                first_weekday = first_day.weekday()
                
                days_to_first_friday = (4 - first_weekday) % 7
                first_friday = first_day + timedelta(days=days_to_first_friday)
                third_friday = first_friday + timedelta(days=14)
                
                if third_friday.month == month and third_friday >= today:
                    expirations.append(third_friday.strftime("%Y-%m-%d"))
            
            return expirations if expirations else None
            
        except Exception as e:
            print(f"Finnhub expiration generation error for {symbol}: {e}")
            return None
    
    def get_options_chain(self, symbol, expiration):
        """Finnhub doesn't provide detailed options chains in free tier"""
        return None
    
    def get_price_history(self, symbol, period_days=90):
        try:
            end_date = int(datetime.now().timestamp())
            start_date = int((datetime.now() - timedelta(days=period_days)).timestamp())
            
            data = self.make_request("stock/candle", {
                "symbol": symbol,
                "resolution": "D",
                "from": start_date,
                "to": end_date
            })
            
            if data and data.get('s') == 'ok':
                df = pd.DataFrame({
                    'Open': data['o'],
                    'High': data['h'],
                    'Low': data['l'],
                    'Close': data['c'],
                    'Volume': data['v']
                })
                df.index = pd.to_datetime(data['t'], unit='s')
                return df
            return None
            
        except Exception as e:
            print(f"Finnhub history error for {symbol}: {e}")
            return None

class DataSourceManager:
    """Manages multiple data sources with priority ordering"""
    
    def __init__(self):
        # PRIORITY ORDER: Yahoo Finance -> Finnhub
        self.sources = [
            YahooFinanceSource(),
            FinnhubSource(FINNHUB_API_KEY)
        ]
        
        print(f"Initialized {len(self.sources)} data sources in priority order:")
        for i, source in enumerate(self.sources, 1):
            print(f"  {i}. {source.name}")
    
    def get_current_price(self, symbol):
        for source in self.sources:
            try:
                price = source.get_current_price(symbol)
                if price is not None and price > 0:
                    print(f"Got price ${price:.2f} for {symbol} from {source.name}")
                    return price
            except Exception as e:
                continue
        print(f"No price data available for {symbol}")
        return None
    
    def get_options_expirations(self, symbol):
        for source in self.sources:
            try:
                expirations = source.get_options_expirations(symbol)
                if expirations and len(expirations) > 0:
                    print(f"Got {len(expirations)} expirations for {symbol} from {source.name}")
                    return expirations
            except Exception as e:
                continue
        print(f"No options expirations available for {symbol}")
        return None
    
    def get_options_chain(self, symbol, expiration):
        for source in self.sources:
            try:
                chain = source.get_options_chain(symbol, expiration)
                if chain and ('calls' in chain or 'puts' in chain):
                    calls_count = len(chain.get('calls', []))
                    puts_count = len(chain.get('puts', []))
                    if calls_count > 0 or puts_count > 0:
                        print(f"Got options chain for {symbol} {expiration} from {source.name} ({calls_count} calls, {puts_count} puts)")
                        return chain
            except Exception as e:
                continue
        print(f"No options chain available for {symbol} {expiration}")
        return None
    
    def get_price_history(self, symbol, period_days=90):
        for source in self.sources:
            try:
                history = source.get_price_history(symbol, period_days)
                if history is not None and not history.empty and len(history) >= 30:
                    print(f"Got {len(history)} days of history for {symbol} from {source.name}")
                    return history
            except Exception as e:
                continue
        print(f"No sufficient price history available for {symbol}")
        return None
    
    def cleanup(self):
        """Cleanup connections"""
        pass

# =============================================================================
# EARNINGS CALENDAR FUNCTIONS
# =============================================================================

def get_earnings_calendar(days_ahead=7):
    """Get earnings calendar with multiple fallback sources and save to Excel"""
    print(f"Fetching earnings calendar for next {days_ahead} days...")
    
    tickers = []
    
    # Method 1: Try Finnhub earnings calendar
    try:
        print("Trying Finnhub earnings calendar...")
        finnhub_source = FinnhubSource(FINNHUB_API_KEY)
        
        today = datetime.now().strftime("%Y-%m-%d")
        end_date = (datetime.now() + timedelta(days=days_ahead)).strftime("%Y-%m-%d")
        
        earnings_data = finnhub_source.make_request("calendar/earnings", {
            "from": today,
            "to": end_date
        })
        
        if earnings_data and 'earningsCalendar' in earnings_data:
            for item in earnings_data['earningsCalendar'][:30]:  # Limit to 30
                if 'symbol' in item and 'date' in item:
                    tickers.append({
                        'ticker': item['symbol'],
                        'earnings_date': item['date'],
                        'company': item.get('symbol', item['symbol'])
                    })
                    
    except Exception as e:
        print(f"Finnhub earnings calendar failed: {e}")
    
    # Method 2: Try Yahoo Finance earnings calendar
    if not tickers:
        try:
            print("Trying Yahoo Finance earnings calendar...")
            base_url = "https://finance.yahoo.com/calendar/earnings"
            
            for day_offset in range(days_ahead):
                target_date = datetime.now() + timedelta(days=day_offset)
                date_str = target_date.strftime("%Y-%m-%d")
                
                try:
                    url = f"{base_url}?day={date_str}"
                    headers = {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
                    }
                    
                    response = requests.get(url, headers=headers, timeout=10)
                    if response.status_code == 200:
                        soup = BeautifulSoup(response.content, 'html.parser')
                        
                        table = soup.find('table', {'data-test': 'calendar-table'}) or soup.find('table')
                        if table:
                            rows = table.find_all('tr')[1:]  # Skip header
                            
                            for row in rows[:15]:  # Limit to top 15 per day
                                cells = row.find_all('td')
                                if len(cells) >= 1:
                                    ticker_cell = cells[0]
                                    ticker_text = ticker_cell.get_text().strip()
                                    
                                    ticker_match = re.search(r'([A-Z]{1,5})', ticker_text)
                                    if ticker_match:
                                        ticker = ticker_match.group(1)
                                        if len(ticker) <= 5:
                                            tickers.append({
                                                'ticker': ticker,
                                                'earnings_date': date_str,
                                                'company': ticker_text
                                            })
                    
                    time.sleep(1)  # Be respectful
                    
                except Exception as e:
                    print(f"Error fetching {date_str}: {e}")
                    continue
                    
        except Exception as e:
            print(f"Yahoo earnings calendar failed: {e}")
    
    # Method 3: Fallback list of popular stocks
    if not tickers:
        print("Using fallback list of popular earnings stocks...")
        fallback_tickers = [
            'AAPL', 'MSFT', 'GOOGL', 'AMZN', 'TSLA', 'META', 'NVDA', 'NFLX',
            'AMD', 'INTC', 'CRM', 'ADBE', 'ORCL', 'IBM', 'QCOM', 'AVGO',
            'JPM', 'BAC', 'WFC', 'GS', 'MS', 'C', 'V', 'MA', 'PYPL'
        ]
        
        today_str = datetime.now().strftime("%Y-%m-%d")
        for ticker in fallback_tickers[:20]:
            tickers.append({
                'ticker': ticker,
                'earnings_date': today_str,
                'company': ticker
            })
    
    # Remove duplicates
    seen = set()
    unique_tickers = []
    for item in tickers:
        ticker = item['ticker']
        if ticker not in seen:
            seen.add(ticker)
            unique_tickers.append(item)
    
    print(f"Found {len(unique_tickers)} unique tickers for earnings analysis")
    
    # Create and save earnings calendar DataFrame
    earnings_df = pd.DataFrame(unique_tickers)
    
    # Save earnings calendar to Excel
    export_earnings_calendar(earnings_df, days_ahead)
    
    return earnings_df

def export_earnings_calendar(earnings_df, days_ahead):
    """Export earnings calendar to Excel"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"earnings_calendar_{days_ahead}days_{timestamp}.xlsx"
    
    print(f"Exporting earnings calendar to: {filename}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Earnings Calendar ({days_ahead} Days)"
    
    # Add header with date range
    date_range = f"{datetime.now().strftime('%Y-%m-%d')} to {(datetime.now() + timedelta(days=days_ahead)).strftime('%Y-%m-%d')}"
    ws.append([f"EARNINGS CALENDAR: {date_range}"])
    ws.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([""])  # Empty row
    
    # Headers
    headers = ['Ticker', 'Company', 'Earnings Date', 'Day of Week']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for _, row in earnings_df.iterrows():
        earnings_date = row['earnings_date']
        try:
            date_obj = datetime.strptime(earnings_date, "%Y-%m-%d")
            day_of_week = date_obj.strftime("%A")
        except:
            day_of_week = "Unknown"
        
        ws.append([
            row['ticker'],
            row['company'],
            earnings_date,
            day_of_week
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(filename)
    print(f"Earnings calendar saved to: {filename}")

# =============================================================================
# ANALYSIS FUNCTIONS (unchanged)
# =============================================================================

def filter_dates(dates):
    today = datetime.today().date()
    cutoff_date = today + timedelta(days=45)
    
    sorted_dates = sorted(datetime.strptime(date, "%Y-%m-%d").date() for date in dates)

    arr = []
    for i, date in enumerate(sorted_dates):
        if date >= cutoff_date:
            arr = [d.strftime("%Y-%m-%d") for d in sorted_dates[:i+1]]  
            break
    
    if len(arr) > 0:
        if arr[0] == today.strftime("%Y-%m-%d"):
            return arr[1:]
        return arr

    raise ValueError("No date 45 days or more in the future found.")

def yang_zhang(price_data, window=30, trading_periods=252, return_last_only=True):
    """Calculate Yang-Zhang realized volatility"""
    log_ho = (price_data['High'] / price_data['Open']).apply(np.log)
    log_lo = (price_data['Low'] / price_data['Open']).apply(np.log)
    log_co = (price_data['Close'] / price_data['Open']).apply(np.log)
    
    log_oc = (price_data['Open'] / price_data['Close'].shift(1)).apply(np.log)
    log_oc_sq = log_oc**2
    
    log_cc = (price_data['Close'] / price_data['Close'].shift(1)).apply(np.log)
    log_cc_sq = log_cc**2
    
    rs = log_ho * (log_ho - log_co) + log_lo * (log_lo - log_co)
    
    close_vol = log_cc_sq.rolling(
        window=window,
        center=False
    ).sum() * (1.0 / (window - 1.0))

    open_vol = log_oc_sq.rolling(
        window=window,
        center=False
    ).sum() * (1.0 / (window - 1.0))

    window_rs = rs.rolling(
        window=window,
        center=False
    ).sum() * (1.0 / (window - 1.0))

    k = 0.34 / (1.34 + ((window + 1) / (window - 1)) )
    result = (open_vol + k * close_vol + (1 - k) * window_rs).apply(np.sqrt) * np.sqrt(trading_periods)

    if return_last_only:
        return result.iloc[-1]
    else:
        return result.dropna()

def build_term_structure(days, ivs):
    """Build term structure interpolation function"""
    days = np.array(days)
    ivs = np.array(ivs)

    sort_idx = days.argsort()
    days = days[sort_idx]
    ivs = ivs[sort_idx]

    spline = interp1d(days, ivs, kind='linear', fill_value="extrapolate")

    def term_spline(dte):
        if dte < days[0]:  
            return ivs[0]
        elif dte > days[-1]:
            return ivs[-1]
        else:  
            return float(spline(dte))

    return term_spline

def generate_trade_ideas(ticker, underlying_price, options_chains, iv30_rv30, ts_slope_0_45, avg_volume, straddle_price, exp_dates):
    """Generate trade ideas with REALISTIC thresholds"""
    
    trade_ideas = []
    
    # Get basic strike information
    first_exp = exp_dates[0] if exp_dates else "N/A"
    second_exp = exp_dates[1] if len(exp_dates) > 1 else "N/A"
    
    if underlying_price:
        atm_strike = round(underlying_price)
        otm_call_strike = round(underlying_price * 1.05)
        otm_put_strike = round(underlying_price * 0.95)
    else:
        atm_strike = "N/A"
        otm_call_strike = "N/A" 
        otm_put_strike = "N/A"
    
    # HIGH IV + NEGATIVE SLOPE = Volatility Selling Strategies
    if iv30_rv30 >= IV_RV_THRESHOLD and ts_slope_0_45 <= SLOPE_THRESHOLD:
        # Calendar spread as primary strategy
        if len(exp_dates) >= 2:
            trade_ideas.append({
                'strategy': 'Calendar Spread (RECOMMENDED)',
                'rationale': f'IV overpriced ({iv30_rv30:.2f} ratio) + negative slope ({ts_slope_0_45:.4f}) = sell expensive front, buy cheaper back',
                'setup': f'Sell {first_exp} {atm_strike} Call, Buy {second_exp} {atm_strike} Call',
                'expiration': f'{first_exp} / {second_exp}',
                'max_profit': 'Maximum at ATM strike at front expiration',
                'risk': 'Limited to net debit paid',
                'target': 'Stock stays near current price through front expiration',
                'stop_loss': 'Close if stock moves >7% or 50% loss'
            })
        
        trade_ideas.append({
            'strategy': 'Short Straddle',
            'rationale': f'High IV/RV ratio ({iv30_rv30:.2f}) suggests IV crush opportunity',
            'setup': f'Sell {atm_strike} Call + Sell {atm_strike} Put',
            'expiration': first_exp,
            'max_profit': f'${straddle_price:.2f}' if straddle_price else 'Premium collected',
            'risk': 'Unlimited',
            'target': 'Close at 25-50% profit or before earnings',
            'stop_loss': '200% of credit received'
        })
        
        trade_ideas.append({
            'strategy': 'Iron Condor',
            'rationale': 'High IV with range-bound expectation post-earnings',
            'setup': f'Sell {otm_put_strike}P, Buy {otm_put_strike-5}P, Sell {otm_call_strike}C, Buy {otm_call_strike+5}C',
            'expiration': first_exp,
            'max_profit': 'Net credit received',
            'risk': 'Limited to spread width minus credit',
            'target': 'Stock stays between short strikes',
            'stop_loss': '50% of max loss'
        })
    
    # NEGATIVE SLOPE BUT LOWER IV/RV = Calendar spreads
    elif ts_slope_0_45 <= SLOPE_THRESHOLD and iv30_rv30 < IV_RV_THRESHOLD:
        if len(exp_dates) >= 2:
            trade_ideas.append({
                'strategy': 'Calendar Spread',
                'rationale': f'Negative slope ({ts_slope_0_45:.4f}) favors selling front month',
                'setup': f'Sell {first_exp} {atm_strike} Call/Put, Buy {second_exp} {atm_strike} Call/Put',
                'expiration': f'{first_exp} / {second_exp}',
                'max_profit': 'At ATM strike at front expiration',
                'risk': 'Limited to net debit paid',
                'target': 'Profit from time decay differential',
                'stop_loss': '50% of debit paid'
            })
    
    # HIGH VOLUME + HIGH IV + POSITIVE SLOPE = Directional plays
    elif avg_volume >= VOLUME_THRESHOLD and iv30_rv30 >= IV_RV_THRESHOLD and ts_slope_0_45 > SLOPE_THRESHOLD:
        trade_ideas.append({
            'strategy': 'Long Strangle',
            'rationale': f'High IV ({iv30_rv30:.2f}) + positive slope suggests big move expected',
            'setup': f'Buy {otm_call_strike} Call + Buy {otm_put_strike} Put',
            'expiration': first_exp,
            'max_profit': 'Unlimited',
            'risk': 'Limited to premium paid',
            'target': f'Stock moves beyond ${otm_call_strike} or ${otm_put_strike}',
            'stop_loss': '50% of premium paid'
        })
    
    # LOW VOLUME = Avoid
    elif avg_volume < VOLUME_THRESHOLD:
        trade_ideas.append({
            'strategy': 'AVOID TRADING',
            'rationale': f'Volume {avg_volume:,.0f} too low for reliable pricing',
            'setup': 'Wait for higher volume or choose different underlying',
            'expiration': 'N/A',
            'max_profit': 'N/A',
            'risk': 'High slippage and poor fills',
            'target': 'Find more liquid alternatives',
            'stop_loss': 'N/A'
        })
    
    # DEFAULT CONSERVATIVE APPROACH
    else:
        trade_ideas.append({
            'strategy': 'Conservative Approach',
            'rationale': f'Mixed signals: IV/RV={iv30_rv30:.2f}, Slope={ts_slope_0_45:.4f}, Vol={avg_volume:,.0f}',
            'setup': f'Small position in {atm_strike} straddle or paper trade first',
            'expiration': first_exp,
            'max_profit': 'Limited',
            'risk': 'Manage position size carefully',
            'target': 'Quick scalp or wait for better setup',
            'stop_loss': '25% of premium'
        })
    
    return trade_ideas

def analyze_single_ticker(ticker_symbol, earnings_date=None, progress_callback=None):
    """Analyze single ticker with robust multi-source data"""
    
    if progress_callback:
        progress_callback(f"Analyzing {ticker_symbol}...")
    
    # Initialize data manager with priority sources
    data_manager = DataSourceManager()
    
    print(f"\nAnalyzing {ticker_symbol}...")
    
    try:
        ticker_symbol = ticker_symbol.strip().upper()
        
        # Get current price
        underlying_price = data_manager.get_current_price(ticker_symbol)
        if underlying_price is None:
            return {
                'ticker': ticker_symbol,
                'earnings_date': earnings_date,
                'error': "Unable to retrieve current price from any source",
                'recommendation': 'NO PRICE DATA'
            }
        
        # Get options expirations
        exp_dates = data_manager.get_options_expirations(ticker_symbol)
        if not exp_dates:
            return {
                'ticker': ticker_symbol,
                'earnings_date': earnings_date,
                'error': "No options expirations available from any source",
                'recommendation': 'NO OPTIONS'
            }
        
        # Filter dates for 45+ days out
        try:
            filtered_exp_dates = filter_dates(exp_dates)
        except Exception as e:
            return {
                'ticker': ticker_symbol,
                'earnings_date': earnings_date,
                'error': "Not enough future option expiration dates",
                'recommendation': 'INSUFFICIENT DATES'
            }
        
        # Get options chains and calculate ATM IVs
        options_chains = {}
        atm_iv = {}
        straddle = None
        
        for i, exp_date in enumerate(filtered_exp_dates[:3]):  # Limit to 3 for speed
            chain = data_manager.get_options_chain(ticker_symbol, exp_date)
            if chain:
                options_chains[exp_date] = chain
                
                calls = chain['calls']
                puts = chain['puts']
                
                if not calls.empty and not puts.empty:
                    # Find ATM strikes
                    call_diffs = (calls['strike'] - underlying_price).abs()
                    call_idx = call_diffs.idxmin()
                    
                    put_diffs = (puts['strike'] - underlying_price).abs() 
                    put_idx = put_diffs.idxmin()
                    
                    call_iv = calls.loc[call_idx, 'impliedVolatility']
                    put_iv = puts.loc[put_idx, 'impliedVolatility']
                    
                    if not pd.isna(call_iv) and not pd.isna(put_iv) and call_iv > 0 and put_iv > 0:
                        atm_iv_value = (call_iv + put_iv) / 2.0
                        atm_iv[exp_date] = atm_iv_value
                        
                        # Calculate straddle for first expiration
                        if i == 0:
                            call_bid = calls.loc[call_idx, 'bid']
                            call_ask = calls.loc[call_idx, 'ask'] 
                            put_bid = puts.loc[put_idx, 'bid']
                            put_ask = puts.loc[put_idx, 'ask']
                            
                            if (all(x is not None and not pd.isna(x) and x > 0 
                                   for x in [call_bid, call_ask, put_bid, put_ask])):
                                call_mid = (call_bid + call_ask) / 2
                                put_mid = (put_bid + put_ask) / 2
                                straddle = call_mid + put_mid
        
        if not atm_iv:
            return {
                'ticker': ticker_symbol,
                'earnings_date': earnings_date,
                'error': "Could not calculate ATM implied volatility from any source",
                'recommendation': 'NO IV DATA'
            }
        
        # Build term structure
        today = datetime.today().date()
        dtes = []
        ivs = []
        
        for exp_date, iv in atm_iv.items():
            exp_date_obj = datetime.strptime(exp_date, "%Y-%m-%d").date()
            days_to_expiry = (exp_date_obj - today).days
            dtes.append(days_to_expiry)
            ivs.append(iv)
        
        if len(dtes) < 2:
            return {
                'ticker': ticker_symbol,
                'earnings_date': earnings_date,
                'error': "Need at least 2 expiration dates for term structure",
                'recommendation': 'INSUFFICIENT EXPIRATIONS'
            }
        
        # Calculate term structure slope
        term_spline = build_term_structure(dtes, ivs)
        ts_slope_0_45 = (term_spline(45) - term_spline(dtes[0])) / (45 - dtes[0])
        
        # Get price history and calculate realized volatility
        price_history = data_manager.get_price_history(ticker_symbol, 90)
        if price_history is None or len(price_history) < 30:
            return {
                'ticker': ticker_symbol,
                'earnings_date': earnings_date,
                'error': "Insufficient price history for calculations from any source",
                'recommendation': 'NO HISTORY'
            }
        
        # Calculate realized volatility
        rv30 = yang_zhang(price_history)
        if pd.isna(rv30) or rv30 <= 0:
            return {
                'ticker': ticker_symbol,
                'earnings_date': earnings_date,
                'error': "Could not calculate realized volatility",
                'recommendation': 'RV ERROR'
            }
        
        # Calculate key metrics
        iv30 = term_spline(30)
        iv30_rv30 = iv30 / rv30
        
        avg_volume = price_history['Volume'].rolling(30).mean().dropna()
        if avg_volume.empty:
            avg_volume = price_history['Volume'].mean()
        else:
            avg_volume = avg_volume.iloc[-1]
        
        # Generate trade ideas
        trade_ideas = generate_trade_ideas(
            ticker_symbol, underlying_price, options_chains, iv30_rv30,
            ts_slope_0_45, avg_volume, straddle, list(filtered_exp_dates)
        )
        
        # Determine recommendation with REALISTIC THRESHOLDS
        avg_volume_bool = avg_volume >= VOLUME_THRESHOLD    # 500k
        iv30_rv30_bool = iv30_rv30 >= IV_RV_THRESHOLD       # 0.9
        ts_slope_bool = ts_slope_0_45 <= SLOPE_THRESHOLD    # -0.001

        if avg_volume_bool and iv30_rv30_bool and ts_slope_bool:
            recommendation = "RECOMMENDED"
        elif ts_slope_bool and ((avg_volume_bool and not iv30_rv30_bool) or (iv30_rv30_bool and not avg_volume_bool)):
            recommendation = "CONSIDER" 
        else:
            recommendation = "AVOID"
        
        # Debug output
        print(f"  Price: ${underlying_price:.2f}")
        print(f"  Volume: {avg_volume:,.0f} ({'✓' if avg_volume_bool else '✗'})")
        print(f"  IV/RV: {iv30_rv30:.3f} ({'✓' if iv30_rv30_bool else '✗'})")
        print(f"  Slope: {ts_slope_0_45:.5f} ({'✓' if ts_slope_bool else '✗'})")
        print(f"  Result: {recommendation}")
        
        # Return results
        return {
            'ticker': ticker_symbol,
            'earnings_date': earnings_date,
            'recommendation': recommendation,
            'underlying_price': underlying_price,
            'avg_volume': avg_volume,
            'avg_volume_pass': avg_volume_bool,
            'iv30': iv30,
            'rv30': rv30, 
            'iv30_rv30': iv30_rv30,
            'iv30_rv30_pass': iv30_rv30_bool,
            'ts_slope_0_45': ts_slope_0_45,
            'ts_slope_pass': ts_slope_bool,
            'expected_move': f"{(straddle/underlying_price*100):.2f}%" if straddle else None,
            'expected_move_pct': (straddle/underlying_price*100) if straddle else None,
            'straddle_price': straddle,
            'trade_ideas': trade_ideas,
            'error': None
        }
        
    except Exception as e:
        print(f"Error analyzing {ticker_symbol}: {e}")
        return {
            'ticker': ticker_symbol,
            'earnings_date': earnings_date,
            'error': f"Analysis error: {str(e)}",
            'recommendation': 'ERROR'
        }

# =============================================================================
# EXCEL EXPORT
# =============================================================================

def export_to_excel(results, filename=None):
    """Export results to Excel with enhanced formatting"""
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"earnings_scanner_results_{timestamp}.xlsx"
    
    print(f"Creating Excel file: {filename}")
    
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Enhanced summary headers
    summary_headers = [
        'Ticker', 'Earnings Date', 'Recommendation', 'Price', 'Volume', 'Volume Pass',
        'IV30', 'RV30', 'IV/RV Ratio', 'IV/RV Pass', 'Term Slope', 'Slope Pass',
        'Expected Move', 'Primary Strategy', 'Error'
    ]
    
    summary_ws.append(summary_headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add summary data with enhanced metrics
    for result in results:
        primary_strategy = "N/A"
        if result.get('trade_ideas') and len(result['trade_ideas']) > 0:
            primary_strategy = result['trade_ideas'][0]['strategy']
        
        # Color code recommendations
        rec_color = None
        if result['recommendation'] == 'RECOMMENDED':
            rec_color = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        elif result['recommendation'] == 'CONSIDER':
            rec_color = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange
        elif result['recommendation'] == 'AVOID':
            rec_color = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
        
        summary_row = [
            result['ticker'],
            result.get('earnings_date', 'N/A'),
            result['recommendation'],
            f"${result.get('underlying_price', 0):.2f}" if result.get('underlying_price') else 'N/A',
            f"{result.get('avg_volume', 0):,.0f}" if result.get('avg_volume') else 'N/A',
            'PASS' if result.get('avg_volume_pass') else 'FAIL',
            f"{result.get('iv30', 0):.3f}" if result.get('iv30') else 'N/A',
            f"{result.get('rv30', 0):.3f}" if result.get('rv30') else 'N/A',
            f"{result.get('iv30_rv30', 0):.3f}" if result.get('iv30_rv30') else 'N/A',
            'PASS' if result.get('iv30_rv30_pass') else 'FAIL',
            f"{result.get('ts_slope_0_45', 0):.5f}" if result.get('ts_slope_0_45') else 'N/A',
            'PASS' if result.get('ts_slope_pass') else 'FAIL',
            result.get('expected_move', 'N/A'),
            primary_strategy,
            result.get('error', '')
        ]
        
        row_num = summary_ws.max_row + 1
        summary_ws.append(summary_row)
        
        # Apply recommendation color
        if rec_color:
            summary_ws.cell(row=row_num, column=3).fill = rec_color
    
    # Auto-adjust column widths
    for column in summary_ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        summary_ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Create individual ticker sheets
    for result in results:
        if result.get('error') and not result.get('underlying_price'):
            continue
            
        ticker = result['ticker']
        ws = wb.create_sheet(ticker)
        
        # Ticker analysis header
        ws.append([f"ANALYSIS REPORT: {ticker}", ""])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append(["Data Sources:", "Yahoo Finance → Finnhub"])
        ws.append(["", ""])
        
        # Analysis metrics
        analysis_data = [
            ['BASIC INFO', ''],
            ['Ticker Symbol', ticker],
            ['Earnings Date', result.get('earnings_date', 'N/A')],
            ['Current Price', f"${result.get('underlying_price', 0):.2f}"],
            ['Final Recommendation', result['recommendation']],
            ['', ''],
            ['VOLUME ANALYSIS', ''],
            ['30-Day Average Volume', f"{result.get('avg_volume', 0):,.0f}"],
            ['Volume Threshold', f"{VOLUME_THRESHOLD:,.0f}"],
            ['Volume Pass', 'PASS' if result.get('avg_volume_pass') else 'FAIL'],
            ['', ''],
            ['VOLATILITY ANALYSIS', ''],
            ['Implied Volatility (30d)', f"{result.get('iv30', 0):.3f}"],
            ['Realized Volatility (30d)', f"{result.get('rv30', 0):.3f}"],
            ['IV/RV Ratio', f"{result.get('iv30_rv30', 0):.3f}"],
            ['IV/RV Threshold', f"{IV_RV_THRESHOLD:.2f}"],
            ['IV/RV Pass', 'PASS' if result.get('iv30_rv30_pass') else 'FAIL'],
            ['', ''],
            ['TERM STRUCTURE ANALYSIS', ''],
            ['Term Structure Slope', f"{result.get('ts_slope_0_45', 0):.6f}"],
            ['Slope Threshold', f"{SLOPE_THRESHOLD:.3f}"],
            ['Slope Pass', 'PASS' if result.get('ts_slope_pass') else 'FAIL'],
            ['', ''],
            ['EXPECTED MOVE', ''],
            ['Straddle Price', f"${result.get('straddle_price', 0):.2f}" if result.get('straddle_price') else 'N/A'],
            ['Expected Move %', result.get('expected_move', 'N/A')],
        ]
        
        for row_data in analysis_data:
            ws.append(row_data)
        
        # Add trade ideas
        if result.get('trade_ideas'):
            ws.append(['', ''])
            ws.append(['TRADE IDEAS', ''])
            ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
            
            for i, idea in enumerate(result['trade_ideas'], 1):
                ws.append(['', ''])
                ws.append([f'Strategy {i}: {idea["strategy"]}', ''])
                ws.cell(ws.max_row, 1).font = Font(bold=True, color="0000FF")
                
                ws.append(['Rationale', idea['rationale']])
                ws.append(['Setup', idea['setup']])
                ws.append(['Expiration', idea['expiration']])
                ws.append(['Max Profit', idea['max_profit']])
                ws.append(['Risk', idea['risk']])
                ws.append(['Target', idea['target']])
                ws.append(['Stop Loss', idea['stop_loss']])
        
        # Style the sheet
        for row in ws.iter_rows():
            for cell in row:
                if cell.column == 1:  # First column
                    cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
    
    # Save workbook
    wb.save(filename)
    print(f"Excel file saved: {filename}")
    return filename

# =============================================================================
# GUI AND MAIN FUNCTIONS
# =============================================================================

def run_earnings_scan():
    """Main earnings scan function"""
    
    progress_layout = [
        [sg.Text("Simplified Earnings Scanner", font=("Helvetica", 14, "bold"))],
        [sg.Text("Data Sources: Yahoo Finance → Finnhub", font=("Helvetica", 9))],
        [sg.Text("", key="status", size=(70, 1))],
        [sg.ProgressBar(100, orientation='h', size=(60, 20), key='progress')],
        [sg.Text("", key="current_ticker", size=(70, 1))],
        [sg.Button("Cancel")]
    ]
    
    progress_window = sg.Window("Scanning Earnings Calendar", progress_layout, modal=True, finalize=True)
    
    results = []
    cancelled = False
    
    def update_progress(message):
        progress_window["status"].update(message)
        progress_window.refresh()
    
    def update_current_ticker(ticker):
        progress_window["current_ticker"].update(f"Currently analyzing: {ticker}")
        progress_window.refresh()
    
    try:
        # Step 1: Get earnings calendar
        update_progress("Fetching earnings calendar from multiple sources...")
        progress_window["progress"].update(5)
        
        earnings_df = get_earnings_calendar(days_ahead=7)
        total_tickers = len(earnings_df)
        
        update_progress(f"Found {total_tickers} tickers to analyze")
        progress_window["progress"].update(10)
        
        # Step 2: Analyze each ticker
        successful_analyses = 0
        for i, (_, row) in enumerate(earnings_df.iterrows()):
            # Check for cancel
            event, _ = progress_window.read(timeout=1)
            if event == "Cancel":
                cancelled = True
                break
            
            ticker = row['ticker']
            earnings_date = row['earnings_date']
            
            update_current_ticker(ticker)
            
            # Analyze ticker
            result = analyze_single_ticker(ticker, earnings_date, update_current_ticker)
            results.append(result)
            
            if result['recommendation'] in ['RECOMMENDED', 'CONSIDER', 'AVOID']:
                successful_analyses += 1
            
            # Update progress
            progress_pct = 10 + (i + 1) / total_tickers * 80
            progress_window["progress"].update(progress_pct)
            update_progress(f"Completed {i + 1}/{total_tickers} tickers ({successful_analyses} successful)")
        
        if not cancelled:
            # Step 3: Export to Excel
            update_progress("Exporting comprehensive results to Excel...")
            progress_window["progress"].update(95)
            
            filename = export_to_excel(results)
            
            progress_window["progress"].update(100)
            update_progress(f"Scan complete! Results saved to {filename}")
            
            time.sleep(3)
    
    except Exception as e:
        update_progress(f"Error during scan: {str(e)}")
        time.sleep(3)
    
    finally:
        progress_window.close()
    
    if not cancelled and results:
        # Show enhanced summary
        successful = len([r for r in results if r['recommendation'] in ['RECOMMENDED', 'CONSIDER', 'AVOID']])
        recommended = len([r for r in results if r['recommendation'] == 'RECOMMENDED'])
        consider = len([r for r in results if r['recommendation'] == 'CONSIDER'])
        avoid = len([r for r in results if r['recommendation'] == 'AVOID'])
        
        summary_text = f"""
EARNINGS SCAN COMPLETE!

Analysis Summary:
• Total Tickers Processed: {len(results)}
• Successfully Analyzed: {successful}
• Data Source Failures: {len(results) - successful}

Trading Recommendations:
• RECOMMENDED: {recommended}
• CONSIDER: {consider}  
• AVOID: {avoid}

Data Sources Used (Priority Order):
1. Yahoo Finance (primary source)
2. Finnhub API (backup data)

Results exported to Excel with:
✓ Comprehensive summary sheet
✓ Individual analysis for each ticker
✓ Detailed trade ideas and strategies
✓ Reliable free data sources

Earnings Calendar also saved separately.

Adjusted Thresholds Used:
• Volume: >= {VOLUME_THRESHOLD:,} (more inclusive)
• IV/RV Ratio: >= {IV_RV_THRESHOLD} (realistic market conditions)
• Term Slope: <= {SLOPE_THRESHOLD} (less restrictive)

Next Steps:
1. Review RECOMMENDED tickers first
2. Check trade ideas for specific strategies
3. Verify with your broker before executing
        """
        
        sg.popup_scrolled(summary_text, title="Earnings Scan Results", size=(80, 25), font=("Courier", 10))
    
    return results

def show_trade_ideas_window(trade_ideas, ticker, underlying_price):
    """Enhanced trade ideas display"""
    
    if not trade_ideas:
        sg.popup("No trade ideas generated", title="Trade Ideas")
        return
    
    layout = [
        [sg.Text(f"Trade Ideas for {ticker} @ ${underlying_price:.2f}", 
                font=("Helvetica", 14, "bold"), justification="center")],
        [sg.Text(f"Using Thresholds: Vol>={VOLUME_THRESHOLD:,}, IV/RV>={IV_RV_THRESHOLD}, Slope<={SLOPE_THRESHOLD}", 
                font=("Helvetica", 8), justification="center")],
        [sg.Text("Data: Yahoo Finance → Finnhub", 
                font=("Helvetica", 8), justification="center")],
        [sg.Text("_" * 90)],
    ]
    
    for i, idea in enumerate(trade_ideas):
        strategy_color = "blue"
        if "RECOMMENDED" in idea['strategy']:
            strategy_color = "darkgreen"
        elif "AVOID" in idea['strategy']:
            strategy_color = "darkred"
        
        layout.extend([
            [sg.Text(f"Strategy {i+1}: {idea['strategy']}", 
                    font=("Helvetica", 12, "bold"), text_color=strategy_color)],
            [sg.Text(f"Rationale: {idea['rationale']}", size=(85, None), 
                    font=("Helvetica", 9))],
            [sg.Text(f"Setup: {idea['setup']}", text_color="darkgreen")],
            [sg.Text(f"Expiration: {idea['expiration']}")],
            [sg.Text(f"Max Profit: {idea['max_profit']}")],
            [sg.Text(f"Risk: {idea['risk']}")],
            [sg.Text(f"Target: {idea['target']}", text_color="navy")],
            [sg.Text(f"Stop Loss: {idea['stop_loss']}", text_color="darkred")],
            [sg.Text("_" * 90)],
        ])
    
    layout.append([sg.Button("Close", size=(10, 1))])
    
    window = sg.Window(
        f"Trade Ideas - {ticker}", 
        [[sg.Column(layout, scrollable=True, vertical_scroll_only=True, size=(800, 600))]],
        modal=True, finalize=True, resizable=True
    )
    
    while True:
        event, _ = window.read()
        if event in (sg.WINDOW_CLOSED, "Close"):
            break
    
    window.close()

def main_gui():
    """Enhanced main GUI with simplified data sources"""
    
    main_layout = [
        [sg.Text("Simplified Earnings Volatility Calculator", font=("Helvetica", 16, "bold"), justification="center")],
        [sg.Text("Yahoo Finance → Finnhub", font=("Helvetica", 10), justification="center")],
        [sg.Text("Single Stock Analysis:")],
        [sg.Text("Enter Stock Symbol:"), sg.Input(key="stock", size=(20, 1))],
        [sg.Button("Analyze Single Stock", bind_return_key=True)],
        [sg.Text("", key="recommendation", size=(70, 1))],
        [sg.Text("_" * 80)],
        [sg.Text("Earnings Calendar & Analysis:")],
        [sg.Text("View earnings calendar for the next 7 days")],
        [sg.Button("Export Earnings Calendar", size=(25, 1))],
        [sg.Text("Scan all stocks reporting earnings in the next 7 days")],
        [sg.Button("Run Full Earnings Scan", size=(25, 1))],
        [sg.Text("_" * 80)],
        [sg.Button("Exit")]
    ]
    
    window = sg.Window("Simplified Earnings Calculator v4.0", main_layout, size=(650, 500))
    
    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break

        if event == "Analyze Single Stock":
            window["recommendation"].update("")
            stock = values.get("stock", "")
            
            if not stock.strip():
                window["recommendation"].update("Please enter a stock symbol")
                continue

            # Show loading
            loading_layout = [[sg.Text("Analyzing with free data sources...", justification="center")]]
            loading_window = sg.Window("Loading", loading_layout, modal=True, finalize=True, size=(350, 100))

            result_holder = {}

            def worker():
                try:
                    result = analyze_single_ticker(stock.strip().upper())
                    result_holder['result'] = result
                except Exception as e:
                    result_holder['error'] = str(e)

            thread = threading.Thread(target=worker, daemon=True)
            thread.start()

            while thread.is_alive():
                event_load, _ = loading_window.read(timeout=100)
                if event_load == sg.WINDOW_CLOSED:
                    break
            thread.join(timeout=1)

            loading_window.close()

            # Display results
            if 'error' in result_holder:
                window["recommendation"].update(f"Error: {result_holder['error']}")
            elif 'result' in result_holder:
                result = result_holder['result']
                
                if result.get('error'):
                    window["recommendation"].update(f"Error: {result['error']}")
                else:
                    recommendation = result['recommendation']
                    
                    # Color coding
                    if recommendation == "RECOMMENDED":
                        title_color = "#006600"
                    elif recommendation == "CONSIDER":
                        title_color = "#ff9900"
                    else:
                        title_color = "#800000"
                    
                    result_layout = [
                        [sg.Text(f"{recommendation}", text_color=title_color, font=("Helvetica", 16))],
                        [sg.Text(f"Price: ${result.get('underlying_price', 0):.2f}")],
                        [sg.Text(f"Volume: {result.get('avg_volume', 0):,.0f} ({'PASS' if result.get('avg_volume_pass') else 'FAIL'})",
                                text_color="#006600" if result.get('avg_volume_pass') else "#800000")],
                        [sg.Text(f"IV/RV: {result.get('iv30_rv30', 0):.3f} ({'PASS' if result.get('iv30_rv30_pass') else 'FAIL'})",
                                text_color="#006600" if result.get('iv30_rv30_pass') else "#800000")],
                        [sg.Text(f"Slope: {result.get('ts_slope_0_45', 0):.5f} ({'PASS' if result.get('ts_slope_pass') else 'FAIL'})",
                                text_color="#006600" if result.get('ts_slope_pass') else "#800000")],
                        [sg.Text(f"Expected Move: {result.get('expected_move', 'N/A')}", text_color="blue")],
                        [sg.Text("_" * 40)],
                        [sg.Button("View Trade Ideas", size=(15, 1)), sg.Button("OK")]
                    ]
                    
                    result_window = sg.Window("Analysis Results", result_layout, modal=True, finalize=True)
                    while True:
                        event_result, _ = result_window.read()
                        if event_result == "View Trade Ideas":
                            show_trade_ideas_window(result.get('trade_ideas', []), stock.upper(), result.get('underlying_price', 0))
                        elif event_result in (sg.WINDOW_CLOSED, "OK"):
                            break
                    result_window.close()

        elif event == "Export Earnings Calendar":
            try:
                # Get and save earnings calendar
                earnings_df = get_earnings_calendar(days_ahead=7)
                window["recommendation"].update(f"Earnings calendar exported! Found {len(earnings_df)} stocks.")
            except Exception as e:
                window["recommendation"].update(f"Error exporting calendar: {str(e)}")

        elif event == "Run Full Earnings Scan":
            try:
                results = run_earnings_scan()
                if results:
                    successful = len([r for r in results if r['recommendation'] in ['RECOMMENDED', 'CONSIDER', 'AVOID']])
                    window["recommendation"].update(f"Scan completed! {successful} stocks analyzed successfully.")
                else:
                    window["recommendation"].update("Scan completed but no results generated.")
            except Exception as e:
                window["recommendation"].update(f"Error running scan: {str(e)}")
    
    window.close()

def gui():
    main_gui()

if __name__ == "__main__":
    print("Starting Simplified Earnings Calculator...")
    print(f"Data Sources: Yahoo Finance → Finnhub")
    print(f"Realistic thresholds: Vol>={VOLUME_THRESHOLD:,}, IV/RV>={IV_RV_THRESHOLD}, Slope<={SLOPE_THRESHOLD}")
    
    gui()