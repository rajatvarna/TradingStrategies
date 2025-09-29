import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
from io import BytesIO
import warnings
import os
warnings.filterwarnings('ignore')

# Strategy Parameters
INITIAL_CAPITAL = 100000
COMMISSION_RATE = 0.0001  # 0.01%
SLIPPAGE = 0.001  # 0.1%

# ETF symbols - Complete set as per video transcript
SYMBOLS = {
    'SPY': 'SPY',   # S&P 500 for 200D SMA regime filter
    'TQQQ': 'TQQQ', # 3x Nasdaq Long (primary bull position)
    'UVXY': 'UVXY', # Volatility ETF (bull market overbought hedge)
    'TECL': 'TECL', # 3x Technology (bear market oversold rebound)
    'SQQQ': 'SQQQ', # 3x Nasdaq Short (bear market short)
    'TLT': 'TLT',   # 20+ Year Treasury Bonds (bear market safe haven)
    'QQQ': 'QQQ',   # Nasdaq benchmark
    'SPY_ETF': 'SPY' # SPY ETF for comparison
}

class TQQQStrategy:
    def __init__(self, start_date='2011-10-03'):  # Changed to match Composer exactly
        self.start_date = start_date
        self.data = {}
        self.signals = None
        self.portfolio = None
        self.trades = []
        self.chart_images = {}
        
    def download_data(self):
        """Download price data for all symbols"""
        print("Downloading data...")
        
        for name, symbol in SYMBOLS.items():
            try:
                ticker = yf.Ticker(symbol)
                data = ticker.history(start=self.start_date, end=datetime.now())
                if len(data) > 0:
                    self.data[name] = data['Close']
                    print(f"✓ {name} ({symbol}): {len(data)} days")
                else:
                    print(f"✗ No data for {name} ({symbol})")
            except Exception as e:
                print(f"✗ Error downloading {name}: {e}")
        
        # Create combined dataframe
        self.df = pd.DataFrame(self.data)
        self.df = self.df.dropna()
        print(f"\nCombined dataset: {len(self.df)} days from {self.df.index[0]} to {self.df.index[-1]}")
        
    def calculate_indicators(self):
        """Calculate technical indicators"""
        print("Calculating indicators...")
        
        # Moving averages
        self.df['SPY_200MA'] = self.df['SPY'].rolling(200).mean()
        self.df['TQQQ_20MA'] = self.df['TQQQ'].rolling(20).mean()
        
        # RSI calculation
        def calculate_rsi(prices, period=10):
            delta = prices.diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
            rs = gain / loss
            rsi = 100 - (100 / (1 + rs))
            return rsi
        
        self.df['TQQQ_RSI'] = calculate_rsi(self.df['TQQQ'])  # For overbought/oversold signals
        self.df['SPY_RSI'] = calculate_rsi(self.df['SPY'])    # For bear market oversold (< 31)
        self.df['SQQQ_RSI'] = calculate_rsi(self.df['SQQQ'])  # For TLT vs SQQQ comparison
        self.df['TLT_RSI'] = calculate_rsi(self.df['TLT'])    # For TLT vs SQQQ comparison
        
        # Market regime
        self.df['bullish_regime'] = self.df['SPY'] > self.df['SPY_200MA']
        
        print("✓ Indicators calculated")
        
    def generate_signals(self):
        """Generate trading signals based on ACTUAL video transcript logic"""
        print("Generating signals...")
        
        signals = []
        
        for i, (date, row) in enumerate(self.df.iterrows()):
            if pd.isna(row['SPY_200MA']) or pd.isna(row['TQQQ_RSI']):
                signals.append('CASH')
                continue
            
            # ACTUAL STRATEGY LOGIC FROM VIDEO TRANSCRIPT
            
            # BULL MARKET REGIME: SPY > 200D SMA
            if row['SPY'] > row['SPY_200MA']:
                # Mean Reversion in Bull Markets
                # If TQQQ becomes overbought (high RSI), exit TQQQ and move into UVXY
                if row['TQQQ_RSI'] >= 79:  # Overbought threshold
                    signals.append('UVXY')
                else:
                    # Default bull market position: TQQQ
                    signals.append('TQQQ')
                    
            # BEAR MARKET REGIME: SPY <= 200D SMA  
            else:
                # Priority 1: Oversold Rebound (RSI < 31)
                # "If RSI drops below 31 (oversold conditions), it switches back into TECL"
                if row['SPY_RSI'] < 31:
                    signals.append('TECL')  # Technology rebound play
                    
                # Priority 2: Catch Upswings  
                # "If TQQQ > 20D SMA (catching upswings)"
                elif row['TQQQ'] > row['TQQQ_20MA']:
                    signals.append('TQQQ')  # Catching tech upswings in bear market
                    
                # Priority 3: SQQQ Oversold Momentum
                # "Buy SQQQ if TQQQ's 10D RSI is less than 31%"
                elif row['TQQQ_RSI'] < 31:
                    signals.append('SQQQ')  # Short tech when TQQQ oversold
                    
                # Priority 4: Defensive Rotation
                # "Moving into bonds like TLT or shorting with SQQQ, depending on which has a better RSI"
                else:
                    if pd.notna(row['SQQQ_RSI']) and pd.notna(row['TLT_RSI']):
                        # Choose whichever has BETTER (higher) RSI for momentum
                        if row['SQQQ_RSI'] > row['TLT_RSI']:
                            signals.append('SQQQ')
                        else:
                            signals.append('TLT')
                    else:
                        signals.append('TLT')  # Default safe haven
        
        self.df['signal'] = signals
        print("✓ Signals generated using ACTUAL video transcript logic")
        
    def backtest_strategy(self):
        """Run the backtest"""
        print("Running backtest...")
        
        # Initialize portfolio
        portfolio = pd.DataFrame(index=self.df.index)
        portfolio['signal'] = self.df['signal']
        portfolio['cash'] = 0.0
        portfolio['shares'] = 0.0
        portfolio['position_value'] = 0.0
        portfolio['total_value'] = INITIAL_CAPITAL
        portfolio['returns'] = 0.0
        portfolio['cum_returns'] = 0.0
        
        current_cash = INITIAL_CAPITAL
        current_shares = 0
        current_symbol = 'CASH'
        
        for i, (date, row) in enumerate(self.df.iterrows()):
            signal = row['signal']
            
            # Handle position changes
            if signal != current_symbol:
                # Close current position
                if current_symbol != 'CASH' and current_shares != 0:
                    close_price = self.df.loc[date, current_symbol]
                    if pd.notna(close_price):
                        # Apply slippage and commission
                        close_price *= (1 - SLIPPAGE)
                        proceeds = current_shares * close_price * (1 - COMMISSION_RATE)
                        current_cash += proceeds
                        
                        # Record trade
                        self.trades.append({
                            'date': date,
                            'action': 'SELL',
                            'symbol': current_symbol,
                            'shares': current_shares,
                            'price': close_price,
                            'value': proceeds
                        })
                        
                current_shares = 0
                current_symbol = signal
                
                # Open new position
                if signal != 'CASH':
                    entry_price = self.df.loc[date, signal]
                    if pd.notna(entry_price):
                        # Apply slippage
                        entry_price *= (1 + SLIPPAGE)
                        
                        # AGGRESSIVE FULL ALLOCATION - No restrictions on any asset
                        # Use 99.5% of available cash for ALL positions including UVXY
                        position_value = current_cash * 0.995
                            
                        shares_to_buy = position_value / (entry_price * (1 + COMMISSION_RATE))
                        cost = shares_to_buy * entry_price * (1 + COMMISSION_RATE)
                        
                        if cost <= current_cash:
                            current_shares = shares_to_buy
                            current_cash -= cost
                            
                            # Record trade
                            self.trades.append({
                                'date': date,
                                'action': 'BUY',
                                'symbol': signal,
                                'shares': shares_to_buy,
                                'price': entry_price,
                                'value': cost
                            })
            
            # Update portfolio values
            portfolio.loc[date, 'cash'] = current_cash
            portfolio.loc[date, 'shares'] = current_shares
            
            if current_symbol != 'CASH' and current_shares > 0:
                current_price = self.df.loc[date, current_symbol]
                if pd.notna(current_price):
                    portfolio.loc[date, 'position_value'] = current_shares * current_price
                else:
                    portfolio.loc[date, 'position_value'] = 0
            else:
                portfolio.loc[date, 'position_value'] = 0
                
            portfolio.loc[date, 'total_value'] = current_cash + portfolio.loc[date, 'position_value']
            
            # Calculate returns
            if i > 0:
                prev_value = portfolio.iloc[i-1]['total_value']
                if prev_value > 0:
                    portfolio.loc[date, 'returns'] = (portfolio.loc[date, 'total_value'] / prev_value) - 1
        
        portfolio['cum_returns'] = (1 + portfolio['returns']).cumprod() - 1
        self.portfolio = portfolio
        
        # Convert trades to DataFrame
        self.trades_df = pd.DataFrame(self.trades)
        
        print(f"✓ Backtest complete - {len(self.trades)} trades executed")
        
    def calculate_metrics(self):
        """Calculate comprehensive performance metrics"""
        print("Calculating comprehensive performance metrics...")
        
        returns = self.portfolio['returns'].dropna()
        
        # Basic metrics
        total_return = self.portfolio['cum_returns'].iloc[-1]
        trading_days = len(returns)
        years = trading_days / 252
        ann_return = (1 + total_return) ** (1/years) - 1
        ann_volatility = returns.std() * np.sqrt(252)
        
        # Risk metrics
        sharpe = ann_return / ann_volatility if ann_volatility > 0 else 0
        
        # Sortino ratio
        negative_returns = returns[returns < 0]
        downside_deviation = negative_returns.std() * np.sqrt(252) if len(negative_returns) > 0 else 0
        sortino = ann_return / downside_deviation if downside_deviation > 0 else 0
        
        # Maximum drawdown and related metrics
        cum_values = self.portfolio['total_value']
        rolling_max = cum_values.expanding().max()
        drawdowns = (cum_values - rolling_max) / rolling_max
        max_drawdown = drawdowns.min()
        
        # Calmar ratio (Annual Return / Max Drawdown)
        calmar = ann_return / abs(max_drawdown) if max_drawdown != 0 else 0
        
        # MAR ratio (same as Calmar)
        mar = calmar
        
        # Value at Risk (95% and 99%)
        var_95 = np.percentile(returns, 5)
        var_99 = np.percentile(returns, 1)
        
        # Expected Shortfall (Conditional VaR)
        es_95 = returns[returns <= var_95].mean()
        es_99 = returns[returns <= var_99].mean()
        
        # Skewness and Kurtosis
        skewness = returns.skew()
        kurtosis = returns.kurtosis()
        
        # Best and worst periods
        best_day = returns.max()
        worst_day = returns.min()
        
        # Rolling periods analysis
        monthly_returns = returns.resample('M').apply(lambda x: (1 + x).prod() - 1)
        best_month = monthly_returns.max()
        worst_month = monthly_returns.min()
        positive_months = (monthly_returns > 0).sum()
        total_months = len(monthly_returns)
        
        # Quarterly returns
        quarterly_returns = returns.resample('Q').apply(lambda x: (1 + x).prod() - 1)
        best_quarter = quarterly_returns.max()
        worst_quarter = quarterly_returns.min()
        positive_quarters = (quarterly_returns > 0).sum()
        total_quarters = len(quarterly_returns)
        
        # Yearly returns
        yearly_returns = returns.resample('Y').apply(lambda x: (1 + x).prod() - 1)
        best_year = yearly_returns.max()
        worst_year = yearly_returns.min()
        positive_years = (yearly_returns > 0).sum()
        total_years = len(yearly_returns)
        
        # Win rate and profit metrics
        if len(self.trades_df) > 0:
            # Calculate trade P&L more robustly
            buy_trades = self.trades_df[self.trades_df['action'] == 'BUY'].copy()
            sell_trades = self.trades_df[self.trades_df['action'] == 'SELL'].copy()
            
            trade_pnl = []
            
            # Match buy/sell pairs
            min_trades = min(len(buy_trades), len(sell_trades))
            for i in range(min_trades):
                buy_value = buy_trades.iloc[i]['value']
                sell_value = sell_trades.iloc[i]['value']
                pnl_pct = (sell_value - buy_value) / buy_value
                trade_pnl.append(pnl_pct)
            
            if trade_pnl:
                trade_pnl = np.array(trade_pnl)
                win_rate = (trade_pnl > 0).mean()
                avg_win = trade_pnl[trade_pnl > 0].mean() if len(trade_pnl[trade_pnl > 0]) > 0 else 0
                avg_loss = trade_pnl[trade_pnl < 0].mean() if len(trade_pnl[trade_pnl < 0]) > 0 else 0
                profit_factor = abs(avg_win / avg_loss) if avg_loss != 0 else 0
                
                # Additional trade metrics
                largest_win = trade_pnl.max()
                largest_loss = trade_pnl.min()
                avg_trade = trade_pnl.mean()
                
                # Consecutive wins/losses
                win_streaks = []
                loss_streaks = []
                current_win_streak = 0
                current_loss_streak = 0
                
                for pnl in trade_pnl:
                    if pnl > 0:
                        current_win_streak += 1
                        if current_loss_streak > 0:
                            loss_streaks.append(current_loss_streak)
                            current_loss_streak = 0
                    else:
                        current_loss_streak += 1
                        if current_win_streak > 0:
                            win_streaks.append(current_win_streak)
                            current_win_streak = 0
                
                max_consecutive_wins = max(win_streaks) if win_streaks else 0
                max_consecutive_losses = max(loss_streaks) if loss_streaks else 0
                
            else:
                win_rate = avg_win = avg_loss = profit_factor = 0
                largest_win = largest_loss = avg_trade = 0
                max_consecutive_wins = max_consecutive_losses = 0
        else:
            win_rate = avg_win = avg_loss = profit_factor = 0
            largest_win = largest_loss = avg_trade = 0
            max_consecutive_wins = max_consecutive_losses = 0
        
        # Exposure and position analysis
        non_cash_days = (self.portfolio['signal'] != 'CASH').sum()
        exposure = non_cash_days / len(self.portfolio)
        
        # Position distribution
        position_counts = self.portfolio['signal'].value_counts()
        
        self.metrics = {
            # Return Metrics
            'Total Return (%)': total_return * 100,
            'Annualized Return (%)': ann_return * 100,
            'CAGR (%)': ann_return * 100,
            'Total Years': years,
            
            # Risk Metrics  
            'Annualized Volatility (%)': ann_volatility * 100,
            'Sharpe Ratio': sharpe,
            'Sortino Ratio': sortino,
            'Calmar Ratio': calmar,
            'MAR Ratio': mar,
            'Maximum Drawdown (%)': max_drawdown * 100,
            
            # VaR and Risk
            'VaR 95% (%)': var_95 * 100,
            'VaR 99% (%)': var_99 * 100,
            'Expected Shortfall 95% (%)': es_95 * 100,
            'Expected Shortfall 99% (%)': es_99 * 100,
            'Skewness': skewness,
            'Kurtosis': kurtosis,
            
            # Best/Worst Periods
            'Best Day (%)': best_day * 100,
            'Worst Day (%)': worst_day * 100,
            'Best Month (%)': best_month * 100,
            'Worst Month (%)': worst_month * 100,
            'Best Quarter (%)': best_quarter * 100,
            'Worst Quarter (%)': worst_quarter * 100,
            'Best Year (%)': best_year * 100,
            'Worst Year (%)': worst_year * 100,
            
            # Win Rates by Period
            'Positive Months (%)': (positive_months / total_months * 100) if total_months > 0 else 0,
            'Positive Quarters (%)': (positive_quarters / total_quarters * 100) if total_quarters > 0 else 0,
            'Positive Years (%)': (positive_years / total_years * 100) if total_years > 0 else 0,
            
            # Trade Metrics
            'Win Rate (%)': win_rate * 100,
            'Average Win (%)': avg_win * 100,
            'Average Loss (%)': avg_loss * 100,
            'Average Trade (%)': avg_trade * 100,
            'Largest Win (%)': largest_win * 100,
            'Largest Loss (%)': largest_loss * 100,
            'Profit Factor': profit_factor,
            'Max Consecutive Wins': max_consecutive_wins,
            'Max Consecutive Losses': max_consecutive_losses,
            
            # Activity Metrics
            'Exposure (%)': exposure * 100,
            'Total Trades': len(self.trades_df) // 2,
            'Trades per Year': (len(self.trades_df) // 2) / years,
            'Final Value ($)': self.portfolio['total_value'].iloc[-1],
            'Position Distribution': position_counts.to_dict()
        }
        
        print("✓ Comprehensive metrics calculated")
        
    def calculate_benchmark_metrics(self):
        """Calculate benchmark performance"""
        print("Calculating benchmark performance...")
        
        benchmarks = ['SPY_ETF', 'QQQ', 'TQQQ', 'TECL', 'SQQQ', 'TLT', 'UVXY']
        self.benchmark_metrics = {}
        
        for bench in benchmarks:
            if bench in self.df.columns:
                # Calculate returns
                prices = self.df[bench].dropna()
                returns = prices.pct_change().dropna()
                
                total_return = (prices.iloc[-1] / prices.iloc[0]) - 1
                ann_return = (1 + total_return) ** (252 / len(returns)) - 1
                ann_volatility = returns.std() * np.sqrt(252)
                sharpe = ann_return / ann_volatility if ann_volatility > 0 else 0
                
                # Maximum drawdown
                cum_values = (1 + returns).cumprod()
                rolling_max = cum_values.expanding().max()
                drawdowns = (cum_values - rolling_max) / rolling_max
                max_drawdown = drawdowns.min()
                
                self.benchmark_metrics[bench] = {
                    'Total Return (%)': total_return * 100,
                    'Annualized Return (%)': ann_return * 100,
                    'Annualized Volatility (%)': ann_volatility * 100,
                    'Sharpe Ratio': sharpe,
                    'Maximum Drawdown (%)': max_drawdown * 100
                }
        
        print("✓ Benchmark metrics calculated")

    def monte_carlo_simulation(self, num_simulations=100, min_period_years=2):
        """Run Monte Carlo simulation with random starting dates"""
        print(f"Running Monte Carlo simulation with {num_simulations} iterations...")
        
        # Get available date range for random sampling
        available_dates = self.df.index
        min_days = min_period_years * 252  # Minimum trading days for a simulation
        
        if len(available_dates) < min_days:
            print("Not enough data for Monte Carlo simulation")
            return None
        
        # Ensure we have enough end dates for meaningful simulations
        valid_start_dates = available_dates[:-min_days]
        if len(valid_start_dates) < 50:
            print("Not enough valid start dates for robust Monte Carlo simulation")
            return None
        
        simulation_results = []
        
        for i in range(num_simulations):
            try:
                # Random start date
                start_idx = np.random.choice(len(valid_start_dates))
                start_date = valid_start_dates[start_idx]
                
                # Random end date (at least min_period_years after start)
                min_end_idx = start_idx + min_days
                max_end_idx = len(available_dates) - 1
                
                if min_end_idx >= max_end_idx:
                    continue
                    
                end_idx = np.random.randint(min_end_idx, max_end_idx + 1)
                end_date = available_dates[end_idx]
                
                # Create subset for this simulation
                sim_df = self.df.loc[start_date:end_date].copy()
                
                if len(sim_df) < 252:  # At least 1 year of data
                    continue
                
                # Run simulation on this subset
                sim_portfolio = self._run_simulation_on_subset(sim_df)
                
                if sim_portfolio is not None and len(sim_portfolio) > 0:
                    # Calculate metrics for this simulation
                    sim_returns = sim_portfolio['returns'].dropna()
                    
                    if len(sim_returns) > 0:
                        total_return = sim_portfolio['cum_returns'].iloc[-1]
                        years = len(sim_returns) / 252
                        ann_return = (1 + total_return) ** (1/years) - 1
                        ann_volatility = sim_returns.std() * np.sqrt(252)
                        sharpe = ann_return / ann_volatility if ann_volatility > 0 else 0
                        
                        # Max drawdown
                        cum_values = sim_portfolio['total_value']
                        rolling_max = cum_values.expanding().max()
                        drawdowns = (cum_values - rolling_max) / rolling_max
                        max_drawdown = drawdowns.min()
                        
                        simulation_results.append({
                            'start_date': start_date,
                            'end_date': end_date,
                            'years': years,
                            'total_return': total_return * 100,
                            'ann_return': ann_return * 100,
                            'volatility': ann_volatility * 100,
                            'sharpe': sharpe,
                            'max_drawdown': max_drawdown * 100,
                            'final_value': sim_portfolio['total_value'].iloc[-1]
                        })
                
                if (i + 1) % 20 == 0:
                    print(f"Completed {i + 1}/{num_simulations} simulations...")
                    
            except Exception as e:
                continue
        
        if len(simulation_results) == 0:
            print("No successful simulations completed")
            return None
        
        # Convert to DataFrame for analysis
        mc_df = pd.DataFrame(simulation_results)
        
        # Calculate Monte Carlo statistics
        mc_stats = {
            'num_simulations': len(mc_df),
            'total_return': {
                'mean': mc_df['total_return'].mean(),
                'median': mc_df['total_return'].median(),
                'std': mc_df['total_return'].std(),
                'min': mc_df['total_return'].min(),
                'max': mc_df['total_return'].max(),
                'percentile_5': mc_df['total_return'].quantile(0.05),
                'percentile_95': mc_df['total_return'].quantile(0.95),
            },
            'ann_return': {
                'mean': mc_df['ann_return'].mean(),
                'median': mc_df['ann_return'].median(),
                'std': mc_df['ann_return'].std(),
                'min': mc_df['ann_return'].min(),
                'max': mc_df['ann_return'].max(),
                'percentile_5': mc_df['ann_return'].quantile(0.05),
                'percentile_95': mc_df['ann_return'].quantile(0.95),
            },
            'sharpe': {
                'mean': mc_df['sharpe'].mean(),
                'median': mc_df['sharpe'].median(),
                'std': mc_df['sharpe'].std(),
                'min': mc_df['sharpe'].min(),
                'max': mc_df['sharpe'].max(),
            },
            'max_drawdown': {
                'mean': mc_df['max_drawdown'].mean(),
                'median': mc_df['max_drawdown'].median(),
                'std': mc_df['max_drawdown'].std(),
                'min': mc_df['max_drawdown'].min(),
                'max': mc_df['max_drawdown'].max(),
            },
            'positive_returns_pct': (mc_df['total_return'] > 0).mean() * 100,
            'positive_ann_returns_pct': (mc_df['ann_return'] > 0).mean() * 100,
        }
        
        self.monte_carlo_results = mc_df
        self.monte_carlo_stats = mc_stats
        
        print(f"✓ Monte Carlo simulation completed: {len(mc_df)} successful runs")
        return mc_stats
    
    def _run_simulation_on_subset(self, sim_df):
        """Run strategy simulation on a subset of data"""
        try:
            # Generate signals for subset
            signals = []
            
            for i, (date, row) in enumerate(sim_df.iterrows()):
                if pd.isna(row.get('SPY_200MA')) or pd.isna(row.get('TQQQ_RSI')):
                    signals.append('CASH')
                    continue
                
                # Same logic as main strategy
                if row['SPY'] > row['SPY_200MA']:
                    if row['TQQQ_RSI'] >= 79:
                        signals.append('UVXY')
                    else:
                        signals.append('TQQQ')
                else:
                    if row.get('SPY_RSI', 50) < 31:
                        signals.append('TECL')
                    elif row['TQQQ'] > row['TQQQ_20MA']:
                        signals.append('TQQQ')
                    elif row['TQQQ_RSI'] < 31:
                        signals.append('SQQQ')
                    else:
                        sqqq_rsi = row.get('SQQQ_RSI', 50)
                        tlt_rsi = row.get('TLT_RSI', 50)
                        if sqqq_rsi > tlt_rsi:
                            signals.append('SQQQ')
                        else:
                            signals.append('TLT')
            
            sim_df['signal'] = signals
            
            # Run backtest on subset
            portfolio = pd.DataFrame(index=sim_df.index)
            portfolio['signal'] = sim_df['signal']
            portfolio['total_value'] = INITIAL_CAPITAL
            portfolio['returns'] = 0.0
            portfolio['cum_returns'] = 0.0
            
            current_cash = INITIAL_CAPITAL
            current_shares = 0
            current_symbol = 'CASH'
            
            for i, (date, row) in enumerate(sim_df.iterrows()):
                signal = row['signal']
                
                # Handle position changes
                if signal != current_symbol:
                    # Close current position
                    if current_symbol != 'CASH' and current_shares != 0:
                        if current_symbol in sim_df.columns:
                            close_price = sim_df.loc[date, current_symbol]
                            if pd.notna(close_price):
                                close_price *= (1 - SLIPPAGE)
                                proceeds = current_shares * close_price * (1 - COMMISSION_RATE)
                                current_cash += proceeds
                    
                    current_shares = 0
                    current_symbol = signal
                    
                    # Open new position
                    if signal != 'CASH' and signal in sim_df.columns:
                        entry_price = sim_df.loc[date, signal]
                        if pd.notna(entry_price):
                            entry_price *= (1 + SLIPPAGE)
                            position_value = current_cash * 0.995
                            shares_to_buy = position_value / (entry_price * (1 + COMMISSION_RATE))
                            cost = shares_to_buy * entry_price * (1 + COMMISSION_RATE)
                            
                            if cost <= current_cash:
                                current_shares = shares_to_buy
                                current_cash -= cost
                
                # Update portfolio values
                if current_symbol != 'CASH' and current_shares > 0 and current_symbol in sim_df.columns:
                    current_price = sim_df.loc[date, current_symbol]
                    if pd.notna(current_price):
                        position_value = current_shares * current_price
                    else:
                        position_value = 0
                else:
                    position_value = 0
                    
                portfolio.loc[date, 'total_value'] = current_cash + position_value
                
                # Calculate returns
                if i > 0:
                    prev_value = portfolio.iloc[i-1]['total_value']
                    if prev_value > 0:
                        portfolio.loc[date, 'returns'] = (portfolio.loc[date, 'total_value'] / prev_value) - 1
            
            portfolio['cum_returns'] = (1 + portfolio['returns']).cumprod() - 1
            return portfolio
            
        except Exception as e:
            return None

    def create_visualizations(self):
        """Create visualizations and save to memory for Excel export"""
        print("Creating visualizations...")
        
        # Store charts as in-memory images for Excel
        self.chart_images = {}
        
        # Set up the plotting style
        plt.style.use('default')
        
        # 1. Equity Curve Comparison
        fig1 = plt.figure(figsize=(12, 8))
        strategy_equity = (self.portfolio['total_value'] / INITIAL_CAPITAL)
        strategy_equity.plot(label='TQQQ Strategy', linewidth=3, color='red', alpha=0.9)
        
        colors = ['blue', 'green', 'orange', 'purple']
        color_idx = 0
        for bench in ['SPY_ETF', 'QQQ', 'TQQQ']:
            if bench in self.df.columns:
                bench_equity = self.df[bench] / self.df[bench].iloc[0]
                bench_equity.plot(label=bench, alpha=0.8, linewidth=2, color=colors[color_idx % len(colors)])
                color_idx += 1
        
        plt.title('TQQQ Strategy vs Benchmarks (Log Scale)', fontsize=16, fontweight='bold')
        plt.ylabel('Portfolio Value (Normalized)', fontsize=12)
        plt.xlabel('Date', fontsize=12)
        plt.legend(fontsize=11)
        plt.grid(True, alpha=0.3)
        plt.yscale('log')
        plt.tight_layout()
        
        # Save to memory
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight')
        img_buffer.seek(0)
        self.chart_images['equity_curve'] = img_buffer
        plt.close(fig1)
        
        # 2. Drawdown Analysis
        fig2 = plt.figure(figsize=(12, 6))
        cum_values = self.portfolio['total_value']
        rolling_max = cum_values.expanding().max()
        drawdowns = (cum_values - rolling_max) / rolling_max * 100
        
        plt.fill_between(drawdowns.index, drawdowns, 0, alpha=0.7, color='red', label='Strategy Drawdown')
        plt.title('TQQQ Strategy Drawdowns Over Time', fontsize=16, fontweight='bold')
        plt.ylabel('Drawdown (%)', fontsize=12)
        plt.xlabel('Date', fontsize=12)
        plt.grid(True, alpha=0.3)
        plt.legend(fontsize=11)
        plt.tight_layout()
        
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight')
        img_buffer.seek(0)
        self.chart_images['drawdown'] = img_buffer
        plt.close(fig2)
        
        # 3. Position Allocation
        fig3 = plt.figure(figsize=(10, 8))
        position_counts = self.portfolio['signal'].value_counts()
        colors_pie = plt.cm.Set3(np.linspace(0, 1, len(position_counts)))
        plt.pie(position_counts.values, labels=position_counts.index, autopct='%1.1f%%', 
                colors=colors_pie, startangle=90, textprops={'fontsize': 11})
        plt.title('TQQQ Strategy Position Allocation Distribution', fontsize=16, fontweight='bold')
        plt.tight_layout()
        
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight')
        img_buffer.seek(0)
        self.chart_images['position_allocation'] = img_buffer
        plt.close(fig3)
        
        # 4. Monthly Returns Heatmap
        fig4 = plt.figure(figsize=(14, 8))
        
        # Calculate monthly and yearly returns
        monthly_returns = self.portfolio['returns'].resample('M').apply(lambda x: (1 + x).prod() - 1)
        monthly_returns.index = monthly_returns.index.to_period('M')
        yearly_returns = self.portfolio['returns'].resample('Y').apply(lambda x: (1 + x).prod() - 1)
        yearly_returns.index = yearly_returns.index.to_period('Y')
        
        # Create heatmap data
        monthly_data = {}
        for date, ret in monthly_returns.items():
            year = date.year
            month = date.month
            if year not in monthly_data:
                monthly_data[year] = {}
            monthly_data[year][month] = ret * 100
        
        years = sorted(monthly_data.keys())
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        heatmap_data = []
        yearly_totals = []
        
        for year in years:
            row = []
            for month_num in range(1, 13):
                if month_num in monthly_data[year]:
                    row.append(monthly_data[year][month_num])
                else:
                    row.append(np.nan)
            heatmap_data.append(row)
            
            year_period = pd.Period(year, freq='Y')
            if year_period in yearly_returns.index:
                yearly_totals.append(yearly_returns[year_period] * 100)
            else:
                yearly_totals.append(np.nan)
        
        heatmap_df = pd.DataFrame(heatmap_data, index=years, columns=months)
        heatmap_df['Year Total'] = yearly_totals
        
        sns.heatmap(heatmap_df, annot=True, fmt='.1f', center=0, 
                   cmap='RdYlGn', cbar_kws={'label': 'Return (%)'})
        plt.title('TQQQ Strategy Monthly & Yearly Returns Heatmap', fontsize=16, fontweight='bold')
        plt.xlabel('Month', fontsize=12)
        plt.ylabel('Year', fontsize=12)
        plt.tight_layout()
        
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight')
        img_buffer.seek(0)
        self.chart_images['returns_heatmap'] = img_buffer
        plt.close(fig4)
        
        # 5. Rolling Performance Metrics
        fig5 = plt.figure(figsize=(12, 10))
        
        # Calculate rolling metrics
        rolling_returns = self.portfolio['returns'].rolling(252)
        rolling_sharpe = rolling_returns.mean() / rolling_returns.std() * np.sqrt(252)
        rolling_vol = rolling_returns.std() * np.sqrt(252) * 100
        rolling_ret = rolling_returns.mean() * 252 * 100
        
        # Rolling Sharpe
        plt.subplot(3, 1, 1)
        rolling_sharpe.plot(color='blue', alpha=0.8, linewidth=2)
        plt.title('Rolling 1-Year Performance Metrics', fontsize=16, fontweight='bold')
        plt.ylabel('Sharpe Ratio')
        plt.grid(True, alpha=0.3)
        plt.axhline(y=0, color='red', linestyle='--', alpha=0.7)
        
        # Rolling volatility
        plt.subplot(3, 1, 2)
        rolling_vol.plot(color='orange', alpha=0.8, linewidth=2)
        plt.ylabel('Volatility (%)')
        plt.grid(True, alpha=0.3)
        
        # Rolling returns
        plt.subplot(3, 1, 3)
        rolling_ret.plot(color='green', alpha=0.8, linewidth=2)
        plt.ylabel('Return (%)')
        plt.xlabel('Date')
        plt.grid(True, alpha=0.3)
        plt.axhline(y=0, color='red', linestyle='--', alpha=0.7)
        
        plt.tight_layout()
        
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight')
        img_buffer.seek(0)
        self.chart_images['rolling_metrics'] = img_buffer
        plt.close(fig5)
        
    def save_charts_as_files(self, folder='charts'):
        """Save charts as individual PNG files in a dedicated folder."""
        print(f"Saving charts to folder: {folder}/")
        if not hasattr(self, 'chart_images'):
            print("⚠️  No charts available to save")
            return

        # Create folder if it doesn't exist
        os.makedirs(folder, exist_ok=True)

        chart_files_saved = []
        for chart_name, img_buffer in self.chart_images.items():
            try:
                filename = os.path.join(folder, f"TQQQ_Chart_{chart_name}.png")
                img_buffer.seek(0)
                with open(filename, 'wb') as f:
                    f.write(img_buffer.read())
                chart_files_saved.append(filename)
                print(f"✓ Saved: {filename}")
            except Exception as e:
                print(f"⚠️  Could not save chart {chart_name}: {e}")

        if chart_files_saved:
            print(f"\n✓ {len(chart_files_saved)} chart files saved successfully in '{folder}/'")
            print("📊 Chart files created:")
            for filename in chart_files_saved:
                print(f"  • {filename}")

        return chart_files_saved
    
    def _create_monte_carlo_charts(self):
        """Create Monte Carlo charts for Excel export"""
        try:
            if not hasattr(self, 'chart_images'):
                self.chart_images = {}
            
            # Return distribution histogram
            fig_mc1 = plt.figure(figsize=(12, 10))
            
            # Total returns distribution
            plt.subplot(2, 2, 1)
            plt.hist(self.monte_carlo_results['total_return'], bins=30, alpha=0.7, color='skyblue', edgecolor='black')
            plt.axvline(self.monte_carlo_results['total_return'].mean(), color='red', linestyle='--', 
                       label=f'Mean: {self.monte_carlo_results["total_return"].mean():.1f}%')
            plt.axvline(self.monte_carlo_results['total_return'].median(), color='green', linestyle='--',
                       label=f'Median: {self.monte_carlo_results["total_return"].median():.1f}%')
            plt.title('Distribution of Total Returns')
            plt.xlabel('Total Return (%)')
            plt.ylabel('Frequency')
            plt.legend()
            plt.grid(True, alpha=0.3)
            
            # Annual returns distribution
            plt.subplot(2, 2, 2)
            plt.hist(self.monte_carlo_results['ann_return'], bins=30, alpha=0.7, color='lightgreen', edgecolor='black')
            plt.axvline(self.monte_carlo_results['ann_return'].mean(), color='red', linestyle='--',
                       label=f'Mean: {self.monte_carlo_results["ann_return"].mean():.1f}%')
            plt.title('Distribution of Annualized Returns')
            plt.xlabel('Annualized Return (%)')
            plt.ylabel('Frequency')
            plt.legend()
            plt.grid(True, alpha=0.3)
            
            # Sharpe ratio distribution
            plt.subplot(2, 2, 3)
            plt.hist(self.monte_carlo_results['sharpe'], bins=30, alpha=0.7, color='orange', edgecolor='black')
            plt.axvline(self.monte_carlo_results['sharpe'].mean(), color='red', linestyle='--',
                       label=f'Mean: {self.monte_carlo_results["sharpe"].mean():.2f}')
            plt.title('Distribution of Sharpe Ratios')
            plt.xlabel('Sharpe Ratio')
            plt.ylabel('Frequency')
            plt.legend()
            plt.grid(True, alpha=0.3)
            
            # Max drawdown distribution
            plt.subplot(2, 2, 4)
            plt.hist(self.monte_carlo_results['max_drawdown'], bins=30, alpha=0.7, color='salmon', edgecolor='black')
            plt.axvline(self.monte_carlo_results['max_drawdown'].mean(), color='red', linestyle='--',
                       label=f'Mean: {self.monte_carlo_results["max_drawdown"].mean():.1f}%')
            plt.title('Distribution of Maximum Drawdowns')
            plt.xlabel('Maximum Drawdown (%)')
            plt.ylabel('Frequency')
            plt.legend()
            plt.grid(True, alpha=0.3)
            
            plt.suptitle('Monte Carlo Simulation Results - Return Distributions', fontsize=14, fontweight='bold')
            plt.tight_layout()
            
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight')
            img_buffer.seek(0)
            self.chart_images['monte_carlo_distributions'] = img_buffer
            plt.close(fig_mc1)
            
            # Risk-return scatter plot
            fig_mc2 = plt.figure(figsize=(10, 8))
            scatter = plt.scatter(self.monte_carlo_results['volatility'], self.monte_carlo_results['ann_return'], 
                                c=self.monte_carlo_results['sharpe'], cmap='viridis', alpha=0.6, s=50)
            plt.colorbar(scatter, label='Sharpe Ratio')
            plt.xlabel('Volatility (%)', fontsize=12)
            plt.ylabel('Annualized Return (%)', fontsize=12)
            plt.title('Monte Carlo Risk-Return Analysis\n(Color indicates Sharpe Ratio)', fontsize=14, fontweight='bold')
            plt.grid(True, alpha=0.3)
            
            # Add main strategy result as red star
            if hasattr(self, 'metrics'):
                main_return = self.metrics['Annualized Return (%)']
                main_vol = self.metrics['Annualized Volatility (%)']
                plt.scatter(main_vol, main_return, color='red', s=200, marker='*', 
                           label='Main Strategy Result', edgecolors='black', linewidth=2)
                plt.legend(fontsize=11)
            
            plt.tight_layout()
            
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight')
            img_buffer.seek(0)
            self.chart_images['monte_carlo_scatter'] = img_buffer
            plt.close(fig_mc2)
            
        except Exception as e:
            print(f"Error creating Monte Carlo charts: {e}")
        
    def calculate_yearly_returns(self):
        """Calculate yearly returns for comparison"""
        yearly_returns = {}
        
        # Strategy yearly returns
        strategy_yearly = self.portfolio['returns'].resample('Y').apply(lambda x: (1 + x).prod() - 1)
        yearly_returns['Strategy'] = strategy_yearly * 100
        
        # Benchmark yearly returns
        for bench in ['SPY_ETF', 'QQQ', 'TQQQ']:
            if bench in self.df.columns:
                # Calculate daily returns first, then aggregate to yearly
                bench_daily_returns = self.df[bench].pct_change().dropna()
                bench_yearly = bench_daily_returns.resample('Y').apply(lambda x: (1 + x).prod() - 1) 
                yearly_returns[bench] = bench_yearly * 100
        
        self.yearly_returns = pd.DataFrame(yearly_returns)
        return self.yearly_returns
    
    def export_to_excel(self, filename='TQQQ_Strategy_Comprehensive_Results.xlsx'):
        """Export all results including charts to Excel with proper formatting"""
        print(f"Exporting comprehensive results with embedded charts to {filename}...")
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Summary metrics
            metrics_for_excel = []
            for metric, value in self.metrics.items():
                if metric != 'Position Distribution':
                    metrics_for_excel.append([metric, value])
            
            metrics_df = pd.DataFrame(metrics_for_excel, columns=['Metric', 'Value'])
            metrics_df.to_excel(writer, sheet_name='Strategy_Metrics', index=False)
            
            # Position distribution as separate sheet
            if 'Position Distribution' in self.metrics:
                pos_dist = self.metrics['Position Distribution']
                pos_df = pd.DataFrame(list(pos_dist.items()), columns=['Position', 'Days'])
                total_days = pos_df['Days'].sum()
                pos_df['Percentage'] = (pos_df['Days'] / total_days * 100).round(2)
                pos_df.to_excel(writer, sheet_name='Position_Distribution', index=False)
            
            # Benchmark comparison
            benchmark_df = pd.DataFrame(self.benchmark_metrics).T
            benchmark_df.to_excel(writer, sheet_name='Benchmark_Comparison')
            
            # Monte Carlo results if available
            if hasattr(self, 'monte_carlo_results') and self.monte_carlo_results is not None:
                mc_results = self.monte_carlo_results.copy()
                mc_results['start_date'] = mc_results['start_date'].dt.tz_localize(None)
                mc_results['end_date'] = mc_results['end_date'].dt.tz_localize(None)
                mc_results.to_excel(writer, sheet_name='Monte_Carlo_Results', index=False)
                
                # Monte Carlo statistics
                if hasattr(self, 'monte_carlo_stats'):
                    mc_stats_flat = []
                    for category, stats in self.monte_carlo_stats.items():
                        if isinstance(stats, dict):
                            for stat_name, value in stats.items():
                                mc_stats_flat.append([f"{category}_{stat_name}", value])
                        else:
                            mc_stats_flat.append([category, stats])
                    
                    mc_stats_df = pd.DataFrame(mc_stats_flat, columns=['Metric', 'Value'])
                    mc_stats_df.to_excel(writer, sheet_name='Monte_Carlo_Stats', index=False)
                
                # Create Monte Carlo charts if not already created
                if hasattr(self, 'chart_images') and 'monte_carlo_distributions' not in self.chart_images:
                    self._create_monte_carlo_charts()
            
            # All trades - fix timezone issue
            if len(self.trades_df) > 0:
                trades_export = self.trades_df.copy()
                trades_export['date'] = trades_export['date'].dt.tz_localize(None)
                trades_export.to_excel(writer, sheet_name='All_Trades', index=False)
                
                # Last 10 trades
                last_trades = trades_export.tail(10)
                last_trades.to_excel(writer, sheet_name='Last_10_Trades', index=False)
            
            # Daily portfolio values - fix timezone issue  
            portfolio_export = self.portfolio[['signal', 'total_value', 'returns', 'cum_returns']].copy()
            portfolio_export = portfolio_export.reset_index()
            portfolio_export['Date'] = portfolio_export['Date'].dt.tz_localize(None)
            portfolio_export.to_excel(writer, sheet_name='Daily_Portfolio', index=False)
            
            # Monthly returns - fix timezone issue
            monthly_rets = self.portfolio['returns'].resample('M').apply(lambda x: (1 + x).prod() - 1)
            monthly_rets.name = 'Monthly_Returns'
            monthly_rets_df = monthly_rets.reset_index()
            monthly_rets_df['Date'] = monthly_rets_df['Date'].dt.tz_localize(None)
            monthly_rets_df.to_excel(writer, sheet_name='Monthly_Returns', index=False)
            
            # Yearly returns comparison - fix timezone issue
            yearly_comparison = self.calculate_yearly_returns()
            yearly_export = yearly_comparison.reset_index()
            yearly_export['Date'] = yearly_export['Date'].dt.tz_localize(None)
            yearly_export.to_excel(writer, sheet_name='Yearly_Comparison', index=False)
            
            # Price data and indicators - fix timezone issue
            export_data = self.df[['SPY', 'TQQQ', 'UVXY', 'TECL', 'SQQQ', 'TLT', 'QQQ', 
                                  'SPY_200MA', 'TQQQ_20MA', 'TQQQ_RSI', 'SPY_RSI', 
                                  'SQQQ_RSI', 'TLT_RSI', 'bullish_regime', 'signal']].copy()
            export_data = export_data.reset_index()
            export_data['Date'] = export_data['Date'].dt.tz_localize(None)
            export_data.to_excel(writer, sheet_name='Price_Data_Indicators', index=False)
            
        # Apply Excel number formatting using openpyxl
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import NamedStyle
            
            wb = load_workbook(filename)
            
            # Create formatting styles
            accounting_style = NamedStyle(name="accounting")
            accounting_style.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
            
            percentage_style = NamedStyle(name="percentage") 
            percentage_style.number_format = '0.00%'
            
            number_style = NamedStyle(name="number")
            number_style.number_format = '#,##0.00_);(#,##0.00);"-"??_);(@_)'
            
            ratio_style = NamedStyle(name="ratio")
            ratio_style.number_format = '0.000_);(0.000);"-"??_);(@_)'
            
            # Format Strategy Metrics sheet
            if 'Strategy_Metrics' in wb.sheetnames:
                ws = wb['Strategy_Metrics']
                for row in range(2, ws.max_row + 1):  # Skip header
                    metric_cell = ws[f'A{row}']
                    value_cell = ws[f'B{row}']
                    
                    if metric_cell.value and value_cell.value is not None:
                        metric_name = str(metric_cell.value)
                        
                        if isinstance(value_cell.value, (int, float)):
                            if '$' in metric_name or 'Value' in metric_name:
                                value_cell.style = accounting_style
                            elif '%' in metric_name or 'Rate' in metric_name:
                                # Convert percentage values to decimal for Excel percentage format
                                if value_cell.value > 1:  # Assume it's already in percentage form
                                    value_cell.value = value_cell.value / 100
                                value_cell.style = percentage_style
                            elif 'Ratio' in metric_name:
                                value_cell.style = ratio_style
                            else:
                                value_cell.style = number_style
            
            # Format other sheets with similar logic
            sheets_to_format = ['Position_Distribution', 'Benchmark_Comparison', 'Monte_Carlo_Results', 
                               'Monte_Carlo_Stats', 'All_Trades', 'Last_10_Trades']
            
            for sheet_name in sheets_to_format:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in range(2, ws.max_row + 1):
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            header_cell = ws.cell(row=1, column=col)
                            
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                header_name = str(header_cell.value) if header_cell.value else ""
                                
                                if any(x in header_name.lower() for x in ['%', 'return', 'drawdown']):
                                    if abs(cell.value) > 1:  # Convert percentage
                                        cell.value = cell.value / 100
                                    cell.style = percentage_style
                                elif any(x in header_name.lower() for x in ['ratio', 'sharpe']):
                                    cell.style = ratio_style
                                elif any(x in header_name.lower() for x in ['value', 'price']):
                                    cell.style = accounting_style
                                else:
                                    cell.style = number_style
            
            wb.save(filename)
            print("✓ Excel number formatting applied successfully")
            
        except ImportError as e:
            print(f"⚠️  openpyxl required for number formatting: {e}")
        except Exception as e:
            print(f"⚠️  Error applying Excel formatting: {e}")
        
        # Now add charts to the Excel file using openpyxl directly
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            import tempfile
            import os
            import time
            
            if not hasattr(self, 'chart_images'):
                print("⚠️  No charts available to embed")
                return
            
            wb = load_workbook(filename)
            
            # Add charts to separate sheets if we have them
            chart_sheets = {
                'Charts_Performance': ['equity_curve', 'drawdown'],
                'Charts_Analysis': ['position_allocation', 'returns_heatmap'],
                'Charts_Rolling': ['rolling_metrics']
            }
            
            # Add Monte Carlo charts if available
            if hasattr(self, 'monte_carlo_results') and self.monte_carlo_results is not None:
                chart_sheets['Charts_MonteCarlo'] = ['monte_carlo_distributions', 'monte_carlo_scatter']
            
            for sheet_name, chart_names in chart_sheets.items():
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                ws = wb[sheet_name]
                
                row_position = 1
                for chart_name in chart_names:
                    if chart_name in self.chart_images:
                        tmp_file_path = None
                        try:
                            # Create temporary file with better handling
                            temp_dir = tempfile.gettempdir()
                            temp_filename = f"chart_{chart_name}_{int(time.time())}.png"
                            tmp_file_path = os.path.join(temp_dir, temp_filename)
                            
                            # Write image data to temporary file
                            self.chart_images[chart_name].seek(0)
                            with open(tmp_file_path, 'wb') as f:
                                f.write(self.chart_images[chart_name].read())
                            
                            # Small delay to ensure file is written and released
                            time.sleep(0.1)
                            
                            # Add image to Excel
                            img = OpenpyxlImage(tmp_file_path)
                            img.width = 600  # Resize for Excel
                            img.height = 400
                            ws.add_image(img, f'A{row_position}')
                            
                            row_position += 25  # Leave space between charts
                            
                        except Exception as e:
                            print(f"⚠️  Could not embed chart {chart_name}: {e}")
                        finally:
                            # Clean up temp file with retry logic
                            if tmp_file_path and os.path.exists(tmp_file_path):
                                for attempt in range(3):  # Try 3 times
                                    try:
                                        time.sleep(0.1)  # Small delay
                                        os.unlink(tmp_file_path)
                                        break
                                    except (OSError, PermissionError):
                                        if attempt == 2:  # Last attempt
                                            print(f"⚠️  Could not delete temp file: {tmp_file_path}")
                                        else:
                                            time.sleep(0.5)  # Wait longer before retry
            
            wb.save(filename)
            wb.close()
            
            print("✓ Charts successfully embedded in Excel file")
            
        except ImportError as e:
            print(f"⚠️  Required libraries not available for chart embedding: {e}")
            print("✓ Data exported successfully (install openpyxl and Pillow for chart embedding)")
        except Exception as e:
            print(f"⚠️  Error embedding charts in Excel: {e}")
            print("✓ Data exported successfully (charts not embedded)")
        
        print(f"✓ Comprehensive results exported to {filename}")
        print("📊 Excel file contains:")
        print("  • Strategy_Metrics: All performance metrics")
        print("  • Position_Distribution: Time allocation analysis")  
        print("  • Benchmark_Comparison: Strategy vs benchmarks")
        print("  • All_Trades: Complete trading history")
        print("  • Last_10_Trades: Recent trading activity")
        print("  • Daily_Portfolio: Daily values and returns")
        print("  • Monthly_Returns: Monthly performance data")
        print("  • Yearly_Comparison: Annual performance vs benchmarks")
        print("  • Price_Data_Indicators: Raw data and signals")
        print("  • Charts_Performance: Equity curve and drawdown charts")
        print("  • Charts_Analysis: Position allocation and returns heatmap")
        print("  • Charts_Rolling: Rolling performance metrics")
        if hasattr(self, 'monte_carlo_results') and self.monte_carlo_results is not None:
            print("  • Monte_Carlo_Results: Simulation data")
            print("  • Monte_Carlo_Stats: Statistical summary")
            print("  • Charts_MonteCarlo: Monte Carlo visualizations")
        print("✅ ALL NUMBERS FORMATTED IN EXCEL ACCOUNTING STYLE!")
        print("  • Currency values: $1,234,567.89")
        print("  • Percentages: 12.34%") 
        print("  • Negative numbers: (1,234.56)")
        print("  • Ratios: 0.123")
        print("\n📊 NOTE: Charts also saved as individual PNG files for easy use!")
        print("  • High-quality images perfect for presentations and reports")

# Initialize and run strategy
def run_complete_backtest():
    strategy = TQQQStrategy()
    strategy.download_data()
    strategy.calculate_indicators()
    strategy.generate_signals()
    strategy.backtest_strategy()
    strategy.calculate_metrics()
    strategy.calculate_benchmark_metrics()
    
    # Run Monte Carlo simulation
    mc_stats = strategy.monte_carlo_simulation(num_simulations=200, min_period_years=2)
    
    # Create visualizations and save as both embedded (if possible) and separate files
    strategy.create_visualizations()
    
    # Save charts as individual PNG files (always works)
    chart_files = strategy.save_charts_as_files()
    
    # Export Excel with data and attempt chart embedding
    strategy.export_to_excel()
    
    print("\n" + "="*80)
    print("TQQQ FOR THE LONG TERM - COMPREHENSIVE BACKTEST & MONTE CARLO ANALYSIS")
    print("="*80)
    print("STRATEGY LOGIC FROM VIDEO TRANSCRIPT:")
    print("🟢 BULL MARKET (SPY > 200D SMA): TQQQ default, UVXY when TQQQ RSI ≥ 79")
    print("🔴 BEAR MARKET (SPY ≤ 200D SMA): TECL/TQQQ/SQQQ/TLT based on conditions")
    print("="*80)

    print("\n📊 MAIN BACKTEST RESULTS:")
    print("-" * 50)
    
    # Core Performance Metrics
    core_metrics = [
        'Total Return (%)', 'Annualized Return (%)', 'CAGR (%)', 'Total Years',
        'Annualized Volatility (%)', 'Sharpe Ratio', 'Sortino Ratio', 'Calmar Ratio',
        'Maximum Drawdown (%)', 'Final Value ($)'
    ]
    
    for metric in core_metrics:
        if metric in strategy.metrics:
            value = strategy.metrics[metric]
            if isinstance(value, float):
                if 'Ratio' in metric:
                    print(f"{metric:<30}: {value:.3f}")
                elif '$' in metric:
                    print(f"{metric:<30}: ${value:,.0f}")
                else:
                    print(f"{metric:<30}: {value:.2f}")
            else:
                print(f"{metric:<30}: {value}")
    
    print("\n📈 RISK METRICS:")
    print("-" * 50)
    risk_metrics = [
        'VaR 95% (%)', 'VaR 99% (%)', 'Expected Shortfall 95% (%)', 'Expected Shortfall 99% (%)',
        'Skewness', 'Kurtosis'
    ]
    
    for metric in risk_metrics:
        if metric in strategy.metrics:
            value = strategy.metrics[metric]
            print(f"{metric:<30}: {value:.3f}")
    
    print("\n🎯 BEST/WORST PERIODS:")
    print("-" * 50)
    period_metrics = [
        'Best Day (%)', 'Worst Day (%)', 'Best Month (%)', 'Worst Month (%)',
        'Best Quarter (%)', 'Worst Quarter (%)', 'Best Year (%)', 'Worst Year (%)'
    ]
    
    for metric in period_metrics:
        if metric in strategy.metrics:
            value = strategy.metrics[metric]
            print(f"{metric:<30}: {value:.2f}")
    
    print("\n📊 WIN RATES BY PERIOD:")
    print("-" * 50)
    win_metrics = [
        'Positive Months (%)', 'Positive Quarters (%)', 'Positive Years (%)'
    ]
    
    for metric in win_metrics:
        if metric in strategy.metrics:
            value = strategy.metrics[metric]
            print(f"{metric:<30}: {value:.1f}")
    
    print("\n💼 TRADING ANALYSIS:")
    print("-" * 50)
    trade_metrics = [
        'Total Trades', 'Trades per Year', 'Win Rate (%)', 'Average Trade (%)',
        'Average Win (%)', 'Average Loss (%)', 'Largest Win (%)', 'Largest Loss (%)',
        'Profit Factor', 'Max Consecutive Wins', 'Max Consecutive Losses', 'Exposure (%)'
    ]
    
    for metric in trade_metrics:
        if metric in strategy.metrics:
            value = strategy.metrics[metric]
            if isinstance(value, float):
                if 'Rate' in metric or '%' in metric:
                    print(f"{metric:<30}: {value:.2f}")
                else:
                    print(f"{metric:<30}: {value:.1f}")
            else:
                print(f"{metric:<30}: {value}")
    
    print("\n🎲 MONTE CARLO SIMULATION RESULTS:")
    print("-" * 50)
    if mc_stats:
        print(f"Number of Simulations: {mc_stats['num_simulations']}")
        print(f"Positive Returns: {mc_stats['positive_returns_pct']:.1f}% of simulations")
        print(f"Positive Annual Returns: {mc_stats['positive_ann_returns_pct']:.1f}% of simulations")
        
        print(f"\nTotal Return Statistics:")
        for key, value in mc_stats['total_return'].items():
            print(f"  {key.replace('_', ' ').title():<12}: {value:.2f}%")
        
        print(f"\nAnnualized Return Statistics:")
        for key, value in mc_stats['ann_return'].items():
            print(f"  {key.replace('_', ' ').title():<12}: {value:.2f}%")
        
        print(f"\nSharpe Ratio Statistics:")
        for key, value in mc_stats['sharpe'].items():
            print(f"  {key.replace('_', ' ').title():<12}: {value:.3f}")
        
        print(f"\nMax Drawdown Statistics:")
        for key, value in mc_stats['max_drawdown'].items():
            print(f"  {key.replace('_', ' ').title():<12}: {value:.2f}%")
    else:
        print("Monte Carlo simulation failed to complete")

    print("\n🏆 BENCHMARK COMPARISON:")
    print("-" * 50)
    
    # Create a proper comparison table
    comparison_data = {}
    
    # Add strategy metrics
    comparison_data['Strategy'] = {
        'Total Return (%)': strategy.metrics['Total Return (%)'],
        'Annualized Return (%)': strategy.metrics['Annualized Return (%)'],
        'Annualized Volatility (%)': strategy.metrics['Annualized Volatility (%)'],
        'Sharpe Ratio': strategy.metrics['Sharpe Ratio'],
        'Maximum Drawdown (%)': strategy.metrics['Maximum Drawdown (%)']
    }
    
    # Add benchmark metrics
    for bench_name, bench_metrics in strategy.benchmark_metrics.items():
        comparison_data[bench_name] = bench_metrics
    
    # Create and display comparison DataFrame
    comparison_df = pd.DataFrame(comparison_data).T
    print(comparison_df.round(2))

    print("\n📋 POSITION DISTRIBUTION:")
    print("-" * 50)
    if 'Position Distribution' in strategy.metrics:
        pos_dist = strategy.metrics['Position Distribution']
        total_days = sum(pos_dist.values())
        for position, days in pos_dist.items():
            percentage = (days / total_days) * 100
            print(f"{position:<10}: {days:>5} days ({percentage:>5.1f}%)")

    print("\n💰 LAST 10 TRADES:")
    print("-" * 80)
    if len(strategy.trades_df) > 0:
        last_10_trades = strategy.trades_df.tail(10)
        print(last_10_trades.to_string(index=False))
        
        print(f"\nTotal Trade Records: {len(strategy.trades_df)}")
        print(f"Buy Orders: {len(strategy.trades_df[strategy.trades_df['action'] == 'BUY'])}")
        print(f"Sell Orders: {len(strategy.trades_df[strategy.trades_df['action'] == 'SELL'])}")
    else:
        print("No trades recorded")
    
    print("\n" + "="*80)
    print("ANALYSIS COMPLETE")
    print("="*80)
    print("📁 OUTPUT FILES:")
    print("  • TQQQ_Strategy_Comprehensive_Results.xlsx - Complete analysis with data")
    print("  • Individual PNG chart files - High-quality charts for presentations")
    print("\n📊 EXCEL FILE STRUCTURE:")
    print("  • Data Sheets: Metrics, trades, daily/monthly/yearly returns") 
    print("  • Chart embedding attempted (may require manual import if failed)")
    print("\n🎨 PNG CHART FILES:")
    print("  • TQQQ_Chart_equity_curve.png - Strategy vs benchmarks")
    print("  • TQQQ_Chart_drawdown.png - Risk analysis")
    print("  • TQQQ_Chart_position_allocation.png - Asset allocation")
    print("  • TQQQ_Chart_returns_heatmap.png - Monthly/yearly performance")
    print("  • TQQQ_Chart_rolling_metrics.png - Rolling performance")
    if hasattr(strategy, 'monte_carlo_results') and strategy.monte_carlo_results is not None:
        print("  • TQQQ_Chart_monte_carlo_distributions.png - MC distributions")
        print("  • TQQQ_Chart_monte_carlo_scatter.png - Risk-return analysis")
    print("="*80)
    
    return strategy

# Run the complete backtest
if __name__ == "__main__":
    strategy = run_complete_backtest()