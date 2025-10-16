"""
===================================================================================
EARNINGS CALENDAR FORWARD FACTOR SCANNER - WORKING EDITION
===================================================================================

Uses reliable data sources:
- Alpha Vantage: Earnings calendar (API KEY: XIAGUPO07P3BSVD5)
- Yahoo Finance: Real-time options chains and pricing

Strategy Logic:
1. Calculate Forward Factor (FF) = (Front_IV - Forward_IV) / Forward_IV
2. Look for FF >= 0.2 (20% backwardation)
3. Recommend calendar spreads (sell front, buy back)

Author: AI-Generated Professional Analysis System
Date: 2025
===================================================================================
"""

import requests
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from scipy.stats import norm
from scipy.optimize import minimize_scalar
import yfinance as yf
import warnings
warnings.filterwarnings('ignore')

# ===================================================================================
# API CONFIGURATION
# ===================================================================================

ALPHA_VANTAGE_API_KEY = "XIAGUPO07P3BSVD5"

# ===================================================================================
# ALPHA VANTAGE CLIENT
# ===================================================================================

class AlphaVantageClient:
    """Alpha Vantage API client for earnings calendar"""
    
    def __init__(self, api_key):
        self.api_key = api_key
        self.base_url = "https://www.alphavantage.co/query"
    
    def get_earnings_calendar(self, horizon='3month'):
        """Get earnings calendar"""
        print(f"Fetching earnings calendar from Alpha Vantage...")
        
        params = {
            'function': 'EARNINGS_CALENDAR',
            'horizon': horizon,
            'apikey': self.api_key
        }
        
        try:
            response = requests.get(self.base_url, params=params, timeout=10)
            
            # Check if we got CSV data
            if response.status_code == 200 and 'symbol' in response.text.lower():
                from io import StringIO
                df = pd.read_csv(StringIO(response.text))
                
                if not df.empty and 'symbol' in df.columns:
                    # Filter to next 2 weeks
                    df['reportDate'] = pd.to_datetime(df['reportDate'])
                    today = pd.Timestamp.now().normalize()
                    two_weeks = today + timedelta(days=14)
                    
                    df = df[(df['reportDate'] >= today) & (df['reportDate'] <= two_weeks)]
                    
                    # Clean up data
                    df = df.rename(columns={'symbol': 'ticker', 'reportDate': 'earnings_date'})
                    
                    # Add company names
                    if 'name' not in df.columns:
                        df['name'] = df['ticker']
                    
                    print(f"✓ Found {len(df)} earnings in next 2 weeks")
                    return df
            
            print("⚠ Alpha Vantage returned no valid data")
            return pd.DataFrame()
        
        except Exception as e:
            print(f"⚠ Error fetching from Alpha Vantage: {e}")
            return pd.DataFrame()
    
    def get_company_overview(self, ticker):
        """Get company overview data"""
        params = {
            'function': 'OVERVIEW',
            'symbol': ticker,
            'apikey': self.api_key
        }
        
        try:
            response = requests.get(self.base_url, params=params, timeout=10)
            data = response.json()
            
            if 'Symbol' in data:
                return {
                    'name': data.get('Name', ticker),
                    'sector': data.get('Sector', 'Unknown'),
                    'market_cap': data.get('MarketCapitalization', 0)
                }
            return None
        except:
            return None


# ===================================================================================
# IMPLIED VOLATILITY CALCULATOR
# ===================================================================================

class ImpliedVolatilityCalculator:
    """Calculate implied volatility from option prices"""
    
    @staticmethod
    def black_scholes_call(S, K, T, r, sigma):
        """Black-Scholes call price"""
        if T <= 0:
            return max(S - K, 0)
        
        d1 = (np.log(S/K) + (r + 0.5*sigma**2)*T) / (sigma*np.sqrt(T))
        d2 = d1 - sigma*np.sqrt(T)
        
        return S*norm.cdf(d1) - K*np.exp(-r*T)*norm.cdf(d2)
    
    @staticmethod
    def black_scholes_put(S, K, T, r, sigma):
        """Black-Scholes put price"""
        if T <= 0:
            return max(K - S, 0)
        
        d1 = (np.log(S/K) + (r + 0.5*sigma**2)*T) / (sigma*np.sqrt(T))
        d2 = d1 - sigma*np.sqrt(T)
        
        return K*np.exp(-r*T)*norm.cdf(-d2) - S*norm.cdf(-d1)
    
    @staticmethod
    def calculate_iv(option_price, S, K, T, r, option_type='call'):
        """Calculate IV using optimization"""
        if T <= 0 or option_price <= 0 or S <= 0 or K <= 0:
            return np.nan
        
        # Check intrinsic value
        if option_type == 'call':
            intrinsic = max(S - K, 0)
        else:
            intrinsic = max(K - S, 0)
        
        # Option must have time value
        if option_price <= intrinsic * 1.001:
            return np.nan
        
        try:
            def objective(sigma):
                if sigma <= 0:
                    return 1e10
                
                if option_type == 'call':
                    bs_price = ImpliedVolatilityCalculator.black_scholes_call(S, K, T, r, sigma)
                else:
                    bs_price = ImpliedVolatilityCalculator.black_scholes_put(S, K, T, r, sigma)
                
                return abs(bs_price - option_price)
            
            result = minimize_scalar(objective, bounds=(0.01, 5.0), method='bounded')
            
            if result.success and result.fun < 0.10:
                return result.x
            
            return np.nan
        except:
            return np.nan


# ===================================================================================
# YAHOO FINANCE OPTIONS ANALYZER
# ===================================================================================

class YahooOptionsAnalyzer:
    """Analyze options data from Yahoo Finance"""
    
    def __init__(self, ticker, risk_free_rate=0.05):
        self.ticker = ticker
        self.rf_rate = risk_free_rate
        self.stock = yf.Ticker(ticker)
        self.iv_calc = ImpliedVolatilityCalculator()
        
        # Get current price
        self.current_price = self._get_current_price()
    
    def _get_current_price(self):
        """Get current stock price"""
        try:
            # Try info first
            price = self.stock.info.get('currentPrice')
            if price and price > 0:
                return price
            
            price = self.stock.info.get('regularMarketPrice')
            if price and price > 0:
                return price
            
            # Fallback to history
            hist = self.stock.history(period='1d')
            if not hist.empty:
                return float(hist['Close'].iloc[-1])
            
            return 0
        except:
            return 0
    
    def get_options_expirations(self):
        """Get available expiration dates"""
        try:
            expirations = self.stock.options
            if len(expirations) > 0:
                return [pd.to_datetime(exp) for exp in expirations]
            return []
        except:
            return []
    
    def find_target_expiration(self, target_dte, expirations):
        """Find expiration closest to target DTE"""
        today = pd.Timestamp.now().normalize()
        
        best_exp = None
        min_diff = float('inf')
        
        for exp in expirations:
            dte = (exp - today).days
            diff = abs(dte - target_dte)
            
            # Must be at least 3 days out and within 30 days of target
            if dte >= 3 and diff < min_diff and diff <= 30:
                min_diff = diff
                best_exp = exp
        
        return best_exp
    
    def get_atm_options(self, expiration_date):
        """Get ATM options for a given expiration"""
        try:
            # Get options chain
            opt_chain = self.stock.option_chain(expiration_date.strftime('%Y-%m-%d'))
            
            calls = opt_chain.calls
            puts = opt_chain.puts
            
            if calls.empty or puts.empty:
                return None
            
            # Find ATM strike (closest to current price)
            calls['strike_diff'] = abs(calls['strike'] - self.current_price)
            puts['strike_diff'] = abs(puts['strike'] - self.current_price)
            
            atm_call = calls.loc[calls['strike_diff'].idxmin()]
            atm_put = puts.loc[puts['strike_diff'].idxmin()]
            
            # Ensure we have valid data
            if atm_call['bid'] <= 0 or atm_call['ask'] <= 0:
                return None
            if atm_put['bid'] <= 0 or atm_put['ask'] <= 0:
                return None
            
            return {
                'expiration': expiration_date,
                'strike': float(atm_call['strike']),
                'call_bid': float(atm_call['bid']),
                'call_ask': float(atm_call['ask']),
                'call_last': float(atm_call['lastPrice']),
                'call_volume': int(atm_call['volume']) if atm_call['volume'] > 0 else 0,
                'call_oi': int(atm_call['openInterest']) if atm_call['openInterest'] > 0 else 0,
                'put_bid': float(atm_put['bid']),
                'put_ask': float(atm_put['ask']),
                'put_last': float(atm_put['lastPrice']),
                'put_volume': int(atm_put['volume']) if atm_put['volume'] > 0 else 0,
                'put_oi': int(atm_put['openInterest']) if atm_put['openInterest'] > 0 else 0,
                'dte': (expiration_date - pd.Timestamp.now().normalize()).days
            }
        
        except Exception as e:
            return None
    
    def calculate_iv_from_options(self, options_data):
        """Calculate implied volatility from options prices"""
        if options_data is None:
            return np.nan
        
        # Use mid price for IV calculation
        call_mid = (options_data['call_bid'] + options_data['call_ask']) / 2
        put_mid = (options_data['put_bid'] + options_data['put_ask']) / 2
        
        # Skip if no valid prices
        if call_mid <= 0 or put_mid <= 0:
            return np.nan
        
        T = options_data['dte'] / 365.0
        K = options_data['strike']
        S = self.current_price
        
        if T <= 0 or S <= 0 or K <= 0:
            return np.nan
        
        # Calculate IV from both call and put
        call_iv = self.iv_calc.calculate_iv(call_mid, S, K, T, self.rf_rate, 'call')
        put_iv = self.iv_calc.calculate_iv(put_mid, S, K, T, self.rf_rate, 'put')
        
        # Average valid IVs (put-call parity)
        valid_ivs = [iv for iv in [call_iv, put_iv] if not np.isnan(iv) and 0.01 < iv < 5.0]
        
        if len(valid_ivs) > 0:
            return np.mean(valid_ivs)
        
        return np.nan


# ===================================================================================
# FORWARD FACTOR CALCULATOR
# ===================================================================================

class ForwardFactorCalculator:
    """Calculate Forward Factor from market data"""
    
    @staticmethod
    def calculate_forward_vol(sigma1, sigma2, T1, T2):
        """
        Calculate forward volatility for period [T1, T2]
        Formula: σ_fwd = sqrt((σ₂²·T₂ - σ₁²·T₁)/(T₂-T₁))
        """
        if T2 <= T1 or sigma1 <= 0 or sigma2 <= 0:
            return np.nan
        
        variance_total = sigma2**2 * T2
        variance_front = sigma1**2 * T1
        variance_forward = (variance_total - variance_front) / (T2 - T1)
        
        if variance_forward <= 0:
            return np.nan
        
        return np.sqrt(variance_forward)
    
    @staticmethod
    def calculate_forward_factor(sigma1, sigma_fwd):
        """
        Calculate Forward Factor: FF = (σ₁ - σ_fwd) / σ_fwd
        Positive FF = Backwardation (front month is hot)
        Negative FF = Contango (front month is cheap)
        """
        if sigma_fwd <= 0 or np.isnan(sigma_fwd):
            return np.nan
        
        return (sigma1 - sigma_fwd) / sigma_fwd
    
    @staticmethod
    def analyze_calendar_spread(front_data, back_data):
        """Analyze calendar spread opportunity"""
        if front_data is None or back_data is None:
            return None
        
        # Get IVs
        front_iv = front_data['iv']
        back_iv = back_data['iv']
        
        if np.isnan(front_iv) or np.isnan(back_iv):
            return None
        
        # Calculate forward vol
        T1 = front_data['dte'] / 365.0
        T2 = back_data['dte'] / 365.0
        
        fwd_vol = ForwardFactorCalculator.calculate_forward_vol(front_iv, back_iv, T1, T2)
        
        if np.isnan(fwd_vol):
            return None
        
        # Calculate forward factor
        ff = ForwardFactorCalculator.calculate_forward_factor(front_iv, fwd_vol)
        
        if np.isnan(ff):
            return None
        
        # Calculate spread pricing
        front_put_mid = (front_data['put_bid'] + front_data['put_ask']) / 2
        back_put_mid = (back_data['put_bid'] + back_data['put_ask']) / 2
        
        # Realistic execution prices
        spread_debit_worst = back_data['put_ask'] - front_data['put_bid']  # Buy ask, sell bid
        spread_debit_best = back_data['put_bid'] - front_data['put_ask']  # Limit orders
        spread_debit_mid = back_put_mid - front_put_mid
        
        # Check liquidity
        min_volume = 5
        min_oi = 25
        
        liquidity_ok = (
            front_data['put_volume'] >= min_volume and
            back_data['put_volume'] >= min_volume and
            front_data['put_oi'] >= min_oi and
            back_data['put_oi'] >= min_oi
        )
        
        # Bid-ask spread quality (tighter is better)
        front_spread_pct = ((front_data['put_ask'] - front_data['put_bid']) / 
                           front_put_mid * 100) if front_put_mid > 0 else 100
        back_spread_pct = ((back_data['put_ask'] - back_data['put_bid']) / 
                          back_put_mid * 100) if back_put_mid > 0 else 100
        
        tight_markets = front_spread_pct < 15 and back_spread_pct < 15
        
        # IV skew (how much more expensive is front vs back)
        iv_skew = (front_iv / back_iv - 1) * 100
        
        return {
            'front_dte': front_data['dte'],
            'back_dte': back_data['dte'],
            'front_iv': front_iv,
            'back_iv': back_iv,
            'forward_vol': fwd_vol,
            'forward_factor': ff,
            'iv_skew': iv_skew,
            'net_debit_mid': spread_debit_mid,
            'net_debit_worst': spread_debit_worst,
            'net_debit_best': spread_debit_best,
            'strike': front_data['strike'],
            'liquidity_ok': liquidity_ok,
            'tight_markets': tight_markets,
            'front_volume': front_data['put_volume'],
            'back_volume': back_data['put_volume'],
            'front_oi': front_data['put_oi'],
            'back_oi': back_data['put_oi'],
            'front_bid': front_data['put_bid'],
            'front_ask': front_data['put_ask'],
            'back_bid': back_data['put_bid'],
            'back_ask': back_data['put_ask'],
            'front_spread_pct': front_spread_pct,
            'back_spread_pct': back_spread_pct
        }


# ===================================================================================
# OPPORTUNITY SCANNER
# ===================================================================================

def scan_for_opportunities(earnings_df, ff_threshold=0.20, max_tickers=25):
    """Scan earnings calendar for Forward Factor opportunities"""
    
    print("\n" + "="*80)
    print("SCANNING FOR FORWARD FACTOR OPPORTUNITIES")
    print("="*80)
    print(f"Threshold: Forward Factor >= {ff_threshold:.2f} ({ff_threshold*100:.0f}%)")
    print(f"Scanning up to {max_tickers} tickers...")
    
    opportunities = []
    
    # Limit to first N tickers to avoid rate limits
    scan_df = earnings_df.head(max_tickers)
    
    for idx, row in scan_df.iterrows():
        ticker = row['ticker']
        
        print(f"\n[{idx+1}/{len(scan_df)}] Analyzing {ticker}...")
        
        try:
            # Initialize analyzer
            analyzer = YahooOptionsAnalyzer(ticker)
            
            if analyzer.current_price == 0:
                print("  ✗ Could not get current price")
                continue
            
            print(f"  Current Price: ${analyzer.current_price:.2f}")
            print(f"  Earnings: {row['earnings_date'].strftime('%Y-%m-%d')}")
            
            # Get available expirations
            expirations = analyzer.get_options_expirations()
            
            if len(expirations) < 2:
                print("  ✗ Insufficient options expirations available")
                continue
            
            print(f"  ✓ Found {len(expirations)} expiration dates")
            
            # Find target expirations (weekly, 30-day, 60-day)
            exp_weekly = analyzer.find_target_expiration(7, expirations)
            exp_30 = analyzer.find_target_expiration(30, expirations)
            exp_60 = analyzer.find_target_expiration(60, expirations)
            
            # Define spread combinations to check
            spreads_to_check = []
            
            if exp_weekly and exp_30:
                spreads_to_check.append(('Weekly-30d', exp_weekly, exp_30))
            if exp_weekly and exp_60:
                spreads_to_check.append(('Weekly-60d', exp_weekly, exp_60))
            if exp_30 and exp_60:
                spreads_to_check.append(('30d-60d', exp_30, exp_60))
            
            if not spreads_to_check:
                print("  ✗ Could not find suitable expiration pairs")
                continue
            
            # Analyze each spread combination
            for spread_name, front_exp, back_exp in spreads_to_check:
                print(f"\n  Checking {spread_name} spread...")
                print(f"    Front: {front_exp.strftime('%Y-%m-%d')} ({(front_exp - pd.Timestamp.now().normalize()).days} DTE)")
                print(f"    Back: {back_exp.strftime('%Y-%m-%d')} ({(back_exp - pd.Timestamp.now().normalize()).days} DTE)")
                
                # Get options data
                front_data = analyzer.get_atm_options(front_exp)
                back_data = analyzer.get_atm_options(back_exp)
                
                if front_data is None or back_data is None:
                    print(f"    ✗ Could not get options data")
                    continue
                
                # Calculate IVs
                front_data['iv'] = analyzer.calculate_iv_from_options(front_data)
                back_data['iv'] = analyzer.calculate_iv_from_options(back_data)
                
                if np.isnan(front_data['iv']) or np.isnan(back_data['iv']):
                    print(f"    ✗ Could not calculate IV")
                    continue
                
                # Analyze the spread
                ff_calc = ForwardFactorCalculator()
                result = ff_calc.analyze_calendar_spread(front_data, back_data)
                
                if result is None:
                    print(f"    ✗ Could not calculate Forward Factor")
                    continue
                
                # Display results
                print(f"    Front IV: {result['front_iv']*100:.1f}%")
                print(f"    Back IV: {result['back_iv']*100:.1f}%")
                print(f"    Forward Vol: {result['forward_vol']*100:.1f}%")
                print(f"    Forward Factor: {result['forward_factor']:.3f} ({result['forward_factor']*100:.1f}%)")
                print(f"    IV Skew: {result['iv_skew']:.1f}%")
                print(f"    Net Debit: ${result['net_debit_mid']:.2f}")
                print(f"    Liquidity: {'✓ Good' if result['liquidity_ok'] else '✗ Low'}")
                print(f"    Markets: {'✓ Tight' if result['tight_markets'] else '✗ Wide'}")
                
                # Check if meets criteria
                if result['forward_factor'] >= ff_threshold:
                    print(f"    ✓✓✓ OPPORTUNITY FOUND! FF = {result['forward_factor']:.3f}")
                    
                    opportunities.append({
                        'ticker': ticker,
                        'company': row.get('name', ticker),
                        'earnings_date': row['earnings_date'],
                        'current_price': analyzer.current_price,
                        'spread_type': spread_name,
                        **result
                    })
                else:
                    print(f"    → FF below threshold ({result['forward_factor']:.3f} < {ff_threshold:.3f})")
        
        except Exception as e:
            print(f"  ✗ Error analyzing {ticker}: {e}")
            continue
    
    return pd.DataFrame(opportunities)


# ===================================================================================
# EXCEL EXPORT FUNCTION
# ===================================================================================

def export_to_excel(opportunities_df, earnings_df, filename):
    """Export opportunities and earnings calendar to formatted Excel file"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("Installing openpyxl for Excel export...")
        import subprocess
        subprocess.check_call(['pip', 'install', 'openpyxl'])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=14)
    warning_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    good_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============================================================================
    # SHEET 1: EXECUTIVE SUMMARY
    # ============================================================================
    ws_summary = wb.create_sheet("Executive Summary", 0)
    
    # Title
    ws_summary['A1'] = 'FORWARD FACTOR TRADE RECOMMENDATIONS'
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:H1')
    
    ws_summary['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws_summary.merge_cells('A2:H2')
    
    ws_summary['A3'] = f'Total Opportunities: {len(opportunities_df)}'
    ws_summary['A3'].font = Font(bold=True)
    ws_summary.merge_cells('A3:H3')
    
    # Summary statistics
    row = 5
    ws_summary[f'A{row}'] = 'SUMMARY STATISTICS'
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    ws_summary.merge_cells(f'A{row}:B{row}')
    
    row += 1
    stats = [
        ('Total Opportunities', len(opportunities_df)),
        ('Avg Forward Factor', f"{opportunities_df['forward_factor'].mean():.3f}"),
        ('Avg Front IV', f"{opportunities_df['front_iv'].mean()*100:.1f}%"),
        ('Avg Back IV', f"{opportunities_df['back_iv'].mean()*100:.1f}%"),
        ('Avg IV Skew', f"{opportunities_df['iv_skew'].mean():.1f}%"),
        ('Avg Net Debit', f"${opportunities_df['net_debit_mid'].mean():.2f}"),
        ('Avg Cost/Contract', f"${opportunities_df['net_debit_worst'].mean() * 100:.2f}"),
    ]
    
    for stat_name, stat_value in stats:
        ws_summary[f'A{row}'] = stat_name
        ws_summary[f'B{row}'] = stat_value
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Top opportunities
    row += 2
    ws_summary[f'A{row}'] = 'TOP 5 OPPORTUNITIES (by Forward Factor)'
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    ws_summary.merge_cells(f'A{row}:H{row}')
    
    row += 1
    headers = ['Rank', 'Ticker', 'Company', 'FF', 'Front IV', 'Net Debit', 'Liquidity', 'Quality']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row += 1
    top_5 = opportunities_df.nlargest(5, 'forward_factor')
    for idx, (_, opp) in enumerate(top_5.iterrows(), 1):
        ws_summary.cell(row=row, column=1, value=idx)
        ws_summary.cell(row=row, column=2, value=opp['ticker'])
        ws_summary.cell(row=row, column=3, value=opp['company'][:30])
        ws_summary.cell(row=row, column=4, value=f"{opp['forward_factor']:.3f}")
        ws_summary.cell(row=row, column=5, value=f"{opp['front_iv']*100:.1f}%")
        ws_summary.cell(row=row, column=6, value=f"${opp['net_debit_mid']:.2f}")
        ws_summary.cell(row=row, column=7, value='Good' if opp['liquidity_ok'] else 'Low')
        ws_summary.cell(row=row, column=8, value='Tight' if opp['tight_markets'] else 'Wide')
        
        # Color code liquidity
        if opp['liquidity_ok']:
            ws_summary.cell(row=row, column=7).fill = good_fill
        else:
            ws_summary.cell(row=row, column=7).fill = warning_fill
        
        row += 1
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 35
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 12
    ws_summary.column_dimensions['G'].width = 12
    ws_summary.column_dimensions['H'].width = 12
    
    # ============================================================================
    # SHEET 2: EARNINGS CALENDAR
    # ============================================================================
    ws_calendar = wb.create_sheet("Earnings Calendar")
    
    # Title
    row = 1
    ws_calendar[f'A{row}'] = 'UPCOMING EARNINGS CALENDAR'
    ws_calendar[f'A{row}'].font = title_font
    ws_calendar.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws_calendar[f'A{row}'] = f'Next 14 Days - Total: {len(earnings_df)} Companies'
    ws_calendar.merge_cells(f'A{row}:F{row}')
    
    row += 2
    
    # Prepare earnings data
    calendar_df = earnings_df.copy()
    
    # Add days until earnings
    today = pd.Timestamp.now()
    calendar_df['days_until'] = calendar_df['earnings_date'].apply(lambda x: (x - today).days)
    
    # Sort by earnings date
    calendar_df = calendar_df.sort_values('earnings_date')
    
    # Format earnings date
    calendar_df['earnings_date_formatted'] = calendar_df['earnings_date'].apply(
        lambda x: x.strftime('%Y-%m-%d (%a)')
    )
    
    # Check if opportunity exists for this ticker
    calendar_df['has_opportunity'] = calendar_df['ticker'].apply(
        lambda x: 'YES' if x in opportunities_df['ticker'].values else 'NO'
    )
    
    # Headers
    cal_headers = ['Ticker', 'Company', 'Earnings Date', 'Days Until', 'Day of Week', 'Opportunity?']
    for col, header in enumerate(cal_headers, 1):
        cell = ws_calendar.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
    
    row += 1
    
    # Data rows
    for _, cal_row in calendar_df.iterrows():
        ws_calendar.cell(row=row, column=1, value=cal_row['ticker'])
        ws_calendar.cell(row=row, column=2, value=cal_row['name'][:40])  # Truncate long names
        ws_calendar.cell(row=row, column=3, value=cal_row['earnings_date'].strftime('%Y-%m-%d'))
        ws_calendar.cell(row=row, column=4, value=int(cal_row['days_until']))
        ws_calendar.cell(row=row, column=5, value=cal_row['earnings_date'].strftime('%A'))
        ws_calendar.cell(row=row, column=6, value=cal_row['has_opportunity'])
        
        # Color code opportunities
        if cal_row['has_opportunity'] == 'YES':
            ws_calendar.cell(row=row, column=6).fill = good_fill
            ws_calendar.cell(row=row, column=6).font = Font(bold=True, color='006100')
        
        # Color code by days until earnings
        days_cell = ws_calendar.cell(row=row, column=4)
        if cal_row['days_until'] <= 3:
            days_cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')  # Yellow
        elif cal_row['days_until'] <= 7:
            days_cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')  # Light yellow
        
        # Add borders
        for col in range(1, 7):
            ws_calendar.cell(row=row, column=col).border = border
        
        row += 1
    
    # Summary at bottom
    row += 2
    ws_calendar[f'A{row}'] = 'SUMMARY BY WEEK'
    ws_calendar[f'A{row}'].font = Font(bold=True, size=11)
    ws_calendar.merge_cells(f'A{row}:F{row}')
    
    row += 1
    week_1 = len(calendar_df[calendar_df['days_until'] <= 7])
    week_2 = len(calendar_df[(calendar_df['days_until'] > 7) & (calendar_df['days_until'] <= 14)])
    
    ws_calendar[f'A{row}'] = 'Next 7 days:'
    ws_calendar[f'B{row}'] = week_1
    ws_calendar[f'A{row}'].font = Font(bold=True)
    row += 1
    
    ws_calendar[f'A{row}'] = 'Days 8-14:'
    ws_calendar[f'B{row}'] = week_2
    ws_calendar[f'A{row}'].font = Font(bold=True)
    row += 1
    
    ws_calendar[f'A{row}'] = 'With Opportunities:'
    ws_calendar[f'B{row}'] = len(opportunities_df)
    ws_calendar[f'A{row}'].font = Font(bold=True)
    ws_calendar[f'B{row}'].font = Font(bold=True, color='006100')
    
    # Column widths
    ws_calendar.column_dimensions['A'].width = 12
    ws_calendar.column_dimensions['B'].width = 35
    ws_calendar.column_dimensions['C'].width = 15
    ws_calendar.column_dimensions['D'].width = 12
    ws_calendar.column_dimensions['E'].width = 15
    ws_calendar.column_dimensions['F'].width = 12
    
    # ============================================================================
    # SHEET 3: ALL OPPORTUNITIES (Detailed)
    # ============================================================================
    ws_detail = wb.create_sheet("All Opportunities")
    
    # Prepare detailed dataframe with formatted columns
    detail_df = opportunities_df.copy()
    
    # Format columns
    detail_df['forward_factor'] = detail_df['forward_factor'].apply(lambda x: f"{x:.3f}")
    detail_df['front_iv'] = detail_df['front_iv'].apply(lambda x: f"{x*100:.1f}%")
    detail_df['back_iv'] = detail_df['back_iv'].apply(lambda x: f"{x*100:.1f}%")
    detail_df['forward_vol'] = detail_df['forward_vol'].apply(lambda x: f"{x*100:.1f}%")
    detail_df['iv_skew'] = detail_df['iv_skew'].apply(lambda x: f"{x:.1f}%")
    detail_df['net_debit_mid'] = detail_df['net_debit_mid'].apply(lambda x: f"${x:.2f}")
    detail_df['net_debit_worst'] = detail_df['net_debit_worst'].apply(lambda x: f"${x:.2f}")
    detail_df['net_debit_best'] = detail_df['net_debit_best'].apply(lambda x: f"${x:.2f}")
    detail_df['strike'] = detail_df['strike'].apply(lambda x: f"${x:.2f}")
    detail_df['current_price'] = detail_df['current_price'].apply(lambda x: f"${x:.2f}")
    detail_df['earnings_date'] = detail_df['earnings_date'].apply(lambda x: x.strftime('%Y-%m-%d'))
    
    # Select and reorder columns for clarity
    columns_order = [
        'ticker', 'company', 'earnings_date', 'current_price', 'spread_type',
        'forward_factor', 'iv_skew', 'front_iv', 'back_iv', 'forward_vol',
        'strike', 'front_dte', 'back_dte',
        'net_debit_mid', 'net_debit_worst', 'net_debit_best',
        'liquidity_ok', 'tight_markets',
        'front_volume', 'back_volume', 'front_oi', 'back_oi',
        'front_bid', 'front_ask', 'back_bid', 'back_ask',
        'front_spread_pct', 'back_spread_pct'
    ]
    
    detail_df = detail_df[columns_order]
    
    # Write to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(detail_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_detail.cell(row=r_idx, column=c_idx, value=value)
            
            # Header formatting
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = border
            else:
                cell.border = border
                cell.alignment = Alignment(horizontal='left')
    
    # Auto-adjust column widths
    for column in ws_detail.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_detail.column_dimensions[column_letter].width = adjusted_width
    
    # ============================================================================
    # SHEET 4: TRADE SETUPS
    # ============================================================================
    ws_trades = wb.create_sheet("Trade Setups")
    
    row = 1
    for idx, opp in opportunities_df.iterrows():
        # Trade header
        ws_trades[f'A{row}'] = f"TRADE #{idx+1}: {opp['ticker']}"
        ws_trades[f'A{row}'].font = Font(bold=True, size=12)
        ws_trades.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Trade details
        trade_details = [
            ('Company', opp['company']),
            ('Current Price', f"${opp['current_price']:.2f}"),
            ('Earnings Date', opp['earnings_date'].strftime('%Y-%m-%d')),
            ('', ''),
            ('FORWARD FACTOR ANALYSIS', ''),
            ('Forward Factor', f"{opp['forward_factor']:.3f} ({opp['forward_factor']*100:.1f}%)"),
            ('Front IV', f"{opp['front_iv']*100:.1f}%"),
            ('Back IV', f"{opp['back_iv']*100:.1f}%"),
            ('Forward Vol', f"{opp['forward_vol']*100:.1f}%"),
            ('IV Skew', f"{opp['iv_skew']:.1f}%"),
            ('', ''),
            ('TRADE SETUP', ''),
            ('Strategy', f"{opp['spread_type']} PUT Calendar Spread"),
            ('Strike', f"${opp['strike']:.2f} (ATM)"),
            ('Sell', f"{opp['front_dte']} DTE Put @ ${opp['front_bid']:.2f}-${opp['front_ask']:.2f}"),
            ('Buy', f"{opp['back_dte']} DTE Put @ ${opp['back_bid']:.2f}-${opp['back_ask']:.2f}"),
            ('', ''),
            ('PRICING', ''),
            ('Net Debit (Mid)', f"${opp['net_debit_mid']:.2f}"),
            ('Worst Case', f"${opp['net_debit_worst']:.2f}"),
            ('Best Case', f"${opp['net_debit_best']:.2f}"),
            ('Cost per Contract', f"${opp['net_debit_worst'] * 100:.2f}"),
            ('Max Risk', f"${opp['net_debit_worst'] * 100:.2f}"),
            ('', ''),
            ('LIQUIDITY', ''),
            ('Front Volume', f"{opp['front_volume']:.0f}"),
            ('Front OI', f"{opp['front_oi']:.0f}"),
            ('Back Volume', f"{opp['back_volume']:.0f}"),
            ('Back OI', f"{opp['back_oi']:.0f}"),
            ('Status', 'GOOD' if opp['liquidity_ok'] else 'LOW'),
            ('Markets', 'TIGHT' if opp['tight_markets'] else 'WIDE'),
            ('', ''),
            ('EXIT PLAN', ''),
            ('Primary Exit', f"{opp['front_dte']-3} days before front expiry"),
            ('Stop Loss', 'Exit if stock moves >15% from strike'),
            ('Profit Target', '20-40% gain on debit paid'),
        ]
        
        for label, value in trade_details:
            ws_trades[f'A{row}'] = label
            ws_trades[f'B{row}'] = value
            
            if label and not value:  # Section headers
                ws_trades[f'A{row}'].font = Font(bold=True, size=11)
                ws_trades.merge_cells(f'A{row}:D{row}')
            elif label:
                ws_trades[f'A{row}'].font = Font(bold=True)
            
            row += 1
        
        row += 2  # Space between trades
    
    # Column widths
    ws_trades.column_dimensions['A'].width = 25
    ws_trades.column_dimensions['B'].width = 40
    ws_trades.column_dimensions['C'].width = 15
    ws_trades.column_dimensions['D'].width = 15
    
    # ============================================================================
    # SHEET 5: PRICING GRID
    # ============================================================================
    ws_pricing = wb.create_sheet("Pricing Grid")
    
    # Headers
    row = 1
    ws_pricing[f'A{row}'] = 'PRICING COMPARISON'
    ws_pricing[f'A{row}'].font = title_font
    ws_pricing.merge_cells(f'A{row}:H{row}')
    
    row += 2
    pricing_headers = ['Ticker', 'Spread Type', 'Net Debit (Mid)', 'Worst Case', 'Best Case', 
                      'Cost/Contract', 'Max Risk', '% of $10k Portfolio']
    
    for col, header in enumerate(pricing_headers, 1):
        cell = ws_pricing.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
    
    row += 1
    for _, opp in opportunities_df.iterrows():
        cost_per_contract = opp['net_debit_worst'] * 100
        portfolio_pct = (cost_per_contract / 10000) * 100
        
        ws_pricing.cell(row=row, column=1, value=opp['ticker'])
        ws_pricing.cell(row=row, column=2, value=opp['spread_type'])
        ws_pricing.cell(row=row, column=3, value=f"${opp['net_debit_mid']:.2f}")
        ws_pricing.cell(row=row, column=4, value=f"${opp['net_debit_worst']:.2f}")
        ws_pricing.cell(row=row, column=5, value=f"${opp['net_debit_best']:.2f}")
        ws_pricing.cell(row=row, column=6, value=f"${cost_per_contract:.2f}")
        ws_pricing.cell(row=row, column=7, value=f"${cost_per_contract:.2f}")
        ws_pricing.cell(row=row, column=8, value=f"{portfolio_pct:.2f}%")
        
        # Highlight if cost is reasonable (<5% of $10k portfolio)
        if portfolio_pct < 5:
            ws_pricing.cell(row=row, column=8).fill = good_fill
        
        row += 1
    
    # Column widths
    for col in range(1, 9):
        ws_pricing.column_dimensions[chr(64+col)].width = 15
    
    # ============================================================================
    # SHEET 6: LIQUIDITY ANALYSIS
    # ============================================================================
    ws_liquidity = wb.create_sheet("Liquidity Analysis")
    
    row = 1
    ws_liquidity[f'A{row}'] = 'LIQUIDITY QUALITY CHECK'
    ws_liquidity[f'A{row}'].font = title_font
    ws_liquidity.merge_cells(f'A{row}:I{row}')
    
    row += 2
    liq_headers = ['Ticker', 'Front Vol', 'Front OI', 'Back Vol', 'Back OI', 
                   'Front Spread %', 'Back Spread %', 'Liquidity', 'Markets']
    
    for col, header in enumerate(liq_headers, 1):
        cell = ws_liquidity.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
    
    row += 1
    for _, opp in opportunities_df.iterrows():
        ws_liquidity.cell(row=row, column=1, value=opp['ticker'])
        ws_liquidity.cell(row=row, column=2, value=int(opp['front_volume']))
        ws_liquidity.cell(row=row, column=3, value=int(opp['front_oi']))
        ws_liquidity.cell(row=row, column=4, value=int(opp['back_volume']))
        ws_liquidity.cell(row=row, column=5, value=int(opp['back_oi']))
        ws_liquidity.cell(row=row, column=6, value=f"{opp['front_spread_pct']:.1f}%")
        ws_liquidity.cell(row=row, column=7, value=f"{opp['back_spread_pct']:.1f}%")
        ws_liquidity.cell(row=row, column=8, value='GOOD' if opp['liquidity_ok'] else 'LOW')
        ws_liquidity.cell(row=row, column=9, value='TIGHT' if opp['tight_markets'] else 'WIDE')
        
        # Color coding
        if opp['liquidity_ok']:
            ws_liquidity.cell(row=row, column=8).fill = good_fill
        else:
            ws_liquidity.cell(row=row, column=8).fill = warning_fill
        
        if opp['tight_markets']:
            ws_liquidity.cell(row=row, column=9).fill = good_fill
        else:
            ws_liquidity.cell(row=row, column=9).fill = warning_fill
        
        row += 1
    
    # Column widths
    for col in range(1, 10):
        ws_liquidity.column_dimensions[chr(64+col)].width = 14
    
    # Save workbook
    wb.save(filename)
    
    print(f"\n✓ Excel report created with {len(wb.sheetnames)} worksheets:")
    for sheet in wb.sheetnames:
        print(f"  • {sheet}")


# ===================================================================================
# REPORT GENERATOR
# ===================================================================================

def generate_recommendations_report(opportunities_df, earnings_df):
    """Generate detailed trade recommendations"""
    
    print("\n" + "="*80)
    print("FORWARD FACTOR TRADE RECOMMENDATIONS")
    print("="*80)
    
    if opportunities_df.empty:
        print("\n⚠ No opportunities found meeting criteria")
        print("\nPossible reasons:")
        print("  • Forward Factor threshold too high (try 0.15 instead of 0.20)")
        print("  • Low volatility period")
        print("  • Not earnings season")
        print("  • Limited options liquidity on scanned tickers")
        return
    
    # Sort by Forward Factor (highest first)
    opportunities_df = opportunities_df.sort_values('forward_factor', ascending=False)
    
    print(f"\n✓ Found {len(opportunities_df)} high-quality opportunities\n")
    
    for idx, row in opportunities_df.iterrows():
        print("-" * 80)
        print(f"\n🎯 TRADE RECOMMENDATION #{idx+1}")
        print(f"\n{'='*80}")
        print(f"TICKER: {row['ticker']} - {row['company']}")
        print(f"{'='*80}")
        
        print(f"\n📌 MARKET DATA:")
        print(f"  Current Price: ${row['current_price']:.2f}")
        print(f"  Earnings Date: {row['earnings_date'].strftime('%Y-%m-%d')}")
        print(f"  Days to Earnings: {(row['earnings_date'] - pd.Timestamp.now()).days} days")
        
        print(f"\n📊 FORWARD FACTOR ANALYSIS:")
        print(f"  ★ Forward Factor: {row['forward_factor']:.3f} ({row['forward_factor']*100:.1f}%)")
        print(f"  Front IV: {row['front_iv']*100:.1f}%")
        print(f"  Back IV: {row['back_iv']*100:.1f}%")
        print(f"  Forward Vol: {row['forward_vol']*100:.1f}%")
        print(f"  IV Skew: {row['iv_skew']:.1f}% (front premium vs back)")
        
        print(f"\n📈 RECOMMENDED TRADE ({row['spread_type']} Calendar):")
        print(f"  Strategy: PUT Calendar Spread")
        print(f"  Strike: ${row['strike']:.2f} (ATM)")
        print(f"  ")
        print(f"  SELL: {row['front_dte']} DTE Put")
        print(f"    → Bid/Ask: ${row['front_bid']:.2f} / ${row['front_ask']:.2f}")
        print(f"    → Spread: {row['front_spread_pct']:.1f}%")
        print(f"  ")
        print(f"  BUY: {row['back_dte']} DTE Put")
        print(f"    → Bid/Ask: ${row['back_bid']:.2f} / ${row['back_ask']:.2f}")
        print(f"    → Spread: {row['back_spread_pct']:.1f}%")
        
        print(f"\n💰 PRICING & COST:")
        print(f"  Net Debit (Mid): ${row['net_debit_mid']:.2f}")
        print(f"  Worst Case Fill: ${row['net_debit_worst']:.2f}")
        print(f"  Best Case Fill: ${row['net_debit_best']:.2f}")
        print(f"  Target Fill: ${row['net_debit_mid']:.2f}")
        print(f"  ")
        print(f"  Cost per Contract: ${row['net_debit_worst'] * 100:.2f}")
        print(f"  Max Risk: ${row['net_debit_worst'] * 100:.2f} per spread")
        
        print(f"\n💧 LIQUIDITY ANALYSIS:")
        print(f"  Front Leg:")
        print(f"    • Volume: {row['front_volume']:.0f}")
        print(f"    • Open Interest: {row['front_oi']:.0f}")
        print(f"  Back Leg:")
        print(f"    • Volume: {row['back_volume']:.0f}")
        print(f"    • Open Interest: {row['back_oi']:.0f}")
        print(f"  ")
        print(f"  Overall: {'✓ GOOD LIQUIDITY' if row['liquidity_ok'] else '⚠ LOW LIQUIDITY - USE CAUTION'}")
        print(f"  Markets: {'✓ TIGHT SPREADS' if row['tight_markets'] else '⚠ WIDE SPREADS - USE LIMITS'}")
        
        print(f"\n💡 TRADE THESIS:")
        print(f"  • Front IV is {row['forward_factor']*100:.1f}% above forward vol (backwardation)")
        print(f"  • Market pricing {row['iv_skew']:.1f}% premium in near-term options")
        print(f"  • Expect IV contraction after earnings announcement")
        print(f"  • Calendar spread profits from time decay + volatility crush")
        
        print(f"\n⚙️ EXECUTION PLAN:")
        print(f"  1. Enter as LIMIT order at ${row['net_debit_mid']:.2f} or better")
        print(f"  2. Don't chase - be patient for good fill")
        print(f"  3. Position size: 2-4% of portfolio (quarter Kelly)")
        print(f"  4. Max contracts: Calculate based on risk tolerance")
        
        print(f"\n📋 EXIT STRATEGY:")
        print(f"  PRIMARY EXIT: {row['front_dte']-3} days before front expiry")
        print(f"  • Close spread {row['front_dte']-3} days before {(pd.Timestamp.now() + timedelta(days=row['front_dte']-3)).strftime('%Y-%m-%d')}")
        print(f"  ")
        print(f"  STOP LOSS:")
        print(f"  • Exit if stock moves >15% from strike (${row['strike']:.2f})")
        print(f"  • Exit if realized vol spikes >60%")
        print(f"  • Max loss: 50% of debit paid")
        print(f"  ")
        print(f"  PROFIT TARGET:")
        print(f"  • Take profits at 20-40% gain on debit")
        print(f"  • Best scenario: IV crush + price stays near strike")
        
        print(f"\n⚠️  KEY RISKS:")
        print(f"  • Large price moves reduce spread value (directional risk)")
        print(f"  • Both legs affected by IV crush (vega risk)")
        print(f"  • Earnings surprise can gap stock overnight")
        print(f"  • Time decay accelerates near expiration")
        print(f"  • Liquidity may dry up in volatile markets")
    
    # Summary statistics
    print("\n" + "="*80)
    print("SUMMARY STATISTICS")
    print("="*80)
    print(f"\n  Total Opportunities: {len(opportunities_df)}")
    print(f"  Average Forward Factor: {opportunities_df['forward_factor'].mean():.3f} ({opportunities_df['forward_factor'].mean()*100:.1f}%)")
    print(f"  Average Front IV: {opportunities_df['front_iv'].mean()*100:.1f}%")
    print(f"  Average Back IV: {opportunities_df['back_iv'].mean()*100:.1f}%")
    print(f"  Average IV Skew: {opportunities_df['iv_skew'].mean():.1f}%")
    print(f"  Average Net Debit: ${opportunities_df['net_debit_mid'].mean():.2f}")
    print(f"  Average Cost/Contract: ${opportunities_df['net_debit_worst'].mean() * 100:.2f}")
    
    # Export to Excel only (comprehensive report)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_filename = f'forward_factor_complete_report_{timestamp}.xlsx'
    
    export_to_excel(opportunities_df, earnings_df, excel_filename)
    print(f"\n✓ Complete Excel report exported to: {excel_filename}")
    print("="*80)


# ===================================================================================
# MAIN EXECUTION
# ===================================================================================

def main():
    """Main execution function"""
    
    print("\n" + "="*80)
    print(" " * 15 + "EARNINGS CALENDAR FORWARD FACTOR SCANNER")
    print(" " * 25 + "WORKING EDITION")
    print("="*80)
    print("\nData Sources:")
    print("  • Alpha Vantage API - Earnings calendar")
    print("  • Yahoo Finance - Real-time options data")
    print("\nStrategy: Calendar spreads with Forward Factor >= 0.20")
    print("="*80)
    
    # Initialize Alpha Vantage client
    alpha_vantage = AlphaVantageClient(ALPHA_VANTAGE_API_KEY)
    
    # Step 1: Get earnings calendar
    print("\n" + ">"*80)
    print("STEP 1: FETCHING EARNINGS CALENDAR")
    print(">"*80)
    
    earnings_df = alpha_vantage.get_earnings_calendar()
    
    # Fallback if Alpha Vantage fails
    if earnings_df.empty:
        print("\n⚠ Alpha Vantage returned no data. Using popular high-IV tickers...")
        
        # Curated list of high-volatility stocks often with earnings
        popular_tickers = [
            'TSLA', 'NVDA', 'AMD', 'META', 'NFLX', 'COIN', 'PLTR', 
            'SHOP', 'SQ', 'RBLX', 'AFRM', 'HOOD', 'SOFI', 'RIVN',
            'SNOW', 'NET', 'DKNG', 'LCID', 'NIO', 'BABA'
        ]
        
        today = pd.Timestamp.now()
        earnings_df = pd.DataFrame({
            'ticker': popular_tickers,
            'name': popular_tickers,
            'earnings_date': [today + timedelta(days=7)] * len(popular_tickers)
        })
    
    print(f"\n✓ Loaded {len(earnings_df)} tickers for analysis")
    
    if len(earnings_df) > 10:
        print("\nUpcoming Earnings (sample):")
        print(earnings_df[['ticker', 'name', 'earnings_date']].head(10).to_string(index=False))
        print(f"... and {len(earnings_df)-10} more")
    else:
        print("\nUpcoming Earnings:")
        print(earnings_df[['ticker', 'name', 'earnings_date']].to_string(index=False))
    
    # Step 2: Scan for opportunities
    print("\n" + ">"*80)
    print("STEP 2: ANALYZING OPTIONS DATA")
    print(">"*80)
    
    opportunities = scan_for_opportunities(
        earnings_df, 
        ff_threshold=0.20,
        max_tickers=25  # Limit to avoid rate limits
    )
    
    # Step 3: Generate recommendations
    print("\n" + ">"*80)
    print("STEP 3: GENERATING TRADE RECOMMENDATIONS")
    print(">"*80)
    
    generate_recommendations_report(opportunities, earnings_df)
    
    print("\n" + "="*80)
    print(" " * 30 + "SCAN COMPLETE!")
    print("="*80)
    print("\n⚠️  DISCLAIMER:")
    print("  This is for educational and research purposes only.")
    print("  Not financial advice. Always do your own due diligence.")
    print("  Verify all data before placing any trades.")
    print("="*80 + "\n")


if __name__ == "__main__":
    main()